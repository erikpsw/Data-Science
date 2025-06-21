import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as patches
from collections import defaultdict, deque
import openpyxl
import random
import copy
from typing import List, Dict, Tuple, Set, Optional
import time
import pandas as pd
from Ex1 import Job, DisjunctiveGraph, ScheduleSolver, Operation

MACHINE_LIST = ['C1', 'C2', 'C3', 'C4', 'L1', 'PM1', 'NYM1', 'WYM1', 'RCL1', 'RCL2', 'WX1', 'LX1', 'LX2', 'LX3', 'Z1', 'Z2', 'DHH1', 'XQG1', 'JC1', 'QG1']


class FlexibleOperation:
    """柔性工序类"""
    def __init__(self, job_id: int, operation_id: int, machine_options: List[Tuple[str, int]], product_type: str = None):
        self.job_id = job_id
        self.operation_id = operation_id
        self.machine_options = machine_options  # [(machine_name, processing_time), ...]
        self.selected_machine = None
        self.processing_time = 0
        self.product_type = product_type or f"P{job_id}"
        self.id = f"J{job_id}O{operation_id}"
        
    def select_machine(self, machine_name: str):
        """选择加工机器"""
        for machine, time in self.machine_options:
            if machine == machine_name:
                self.selected_machine = machine
                self.processing_time = time
                return True
        return False
    
    def get_available_machines(self) -> List[str]:
        """获取可用机器列表"""
        return [machine for machine, _ in self.machine_options]
    
    def get_processing_time(self, machine_name: str) -> int:
        """获取在指定机器上的加工时间"""
        for machine, time in self.machine_options:
            if machine == machine_name:
                return time
        return float('inf')
    
    def __str__(self):
        return f"FO(J{self.job_id}O{self.operation_id}, {len(self.machine_options)} machines, product={self.product_type})"


class FlexibleJob:
    """柔性工件类"""
    def __init__(self, job_id: int, operations: List[FlexibleOperation], product_type: str = None):
        self.job_id = job_id
        self.operations = operations
        self.product_type = product_type or f"P{job_id}"
        
        # 更新所有工序的产品类型
        for op in self.operations:
            op.product_type = self.product_type
    
    def __str__(self):
        return f"FlexibleJob(J{self.job_id}, {len(self.operations)} operations, product={self.product_type})"


class FlexibleSchedule:
    """柔性调度"""
    def __init__(self, flexible_jobs: List[FlexibleJob]):
        self.flexible_jobs = flexible_jobs
        self.operations = []
        self.machine_operations = defaultdict(list)
        self.job_operations = defaultdict(list)
        self.all_machines = set()
        
        # 构建操作索引
        for job in flexible_jobs:
            for op in job.operations:
                self.operations.append(op)
                self.job_operations[job.job_id].append(op)
                for machine, _ in op.machine_options:
                    self.all_machines.add(machine)
    
    def get_machine_candidates(self, operation: FlexibleOperation) -> List[str]:
        """获取工序的候选机器"""
        return operation.get_available_machines()


class FlexibleScheduleSolver:
    """柔性调度求解器"""
    def __init__(self, schedule: FlexibleSchedule):
        self.schedule = schedule
        self.best_makespan = float('inf')        
        self.best_solution = None
    
    def calculate_makespan(self, machine_assignment: Dict[str, str]) -> Tuple[int, Dict[str, int]]:
        """
        计算makespan - 通过随机选择工件并为其当前工序分配最早可用时间
        """
        start_times = {}
        job_completion_times = defaultdict(int)  # 记录每个工件当前已完成的工序时间
        machine_completion_times = defaultdict(int)  # 记录每台机器的当前时间
        machine_schedule = defaultdict(list)  # 记录每台机器已安排的工序 [(start_time, end_time, operation), ...]
        
        # 初始化每个工件的当前工序索引
        job_current_operation = {}
        job_operations = defaultdict(list)
        
        # 为每个工件建立工序列表（按工序ID排序）
        for op in self.schedule.operations:
            job_operations[op.job_id].append(op)
        
        # 对每个工件的工序按ID排序
        for job_id in job_operations:
            job_operations[job_id].sort(key=lambda x: x.operation_id)
            job_current_operation[job_id] = 0  # 每个工件从第一个工序开始
        
        # 计算总工序数
        total_operations = len(self.schedule.operations)
        scheduled_count = 0
        
        # 随机调度工序
        while scheduled_count < total_operations:
            # 获取所有还有未调度工序的工件
            available_jobs = []
            for job_id in job_operations:
                if job_current_operation[job_id] < len(job_operations[job_id]):
                    available_jobs.append(job_id)
            
            if not available_jobs:
                break
            
            # 随机选择一个工件
            selected_job = random.choice(available_jobs)
            
            # 获取该工件当前需要加工的工序
            current_op_index = job_current_operation[selected_job]
            current_operation = job_operations[selected_job][current_op_index]
    
            # 获取该工序分配的机器
            assigned_machine = machine_assignment[current_operation.id]
            processing_time = current_operation.get_processing_time(assigned_machine)            
            # print(f"调度工件 {selected_job} 的工序 {current_operation.id} ({processing_time}) (机器分配: {assigned_machine})")
            
            # 工件约束：必须等待同一工件的前序工序完成
            earliest_start_job = job_completion_times[selected_job]
            
            # 机器约束：在满足工件约束的前提下，找到机器上的最早可用时间
            earliest_start_machine = self._find_earliest_available_time(
                machine_schedule[assigned_machine], processing_time, earliest_start_job
            )
            # print(f"{[(d[0],d[1]) for d in machine_schedule[assigned_machine]]}  机器约束: {earliest_start_machine}, 工件约束: {earliest_start_job}")
            
            # 取两者的最大值作为最早开始时间
            earliest_start = max(earliest_start_job, earliest_start_machine)
            completion_time = earliest_start + processing_time
            
            # 记录工序的开始时间
            start_times[current_operation.id] = earliest_start
            # 更新机器调度表
            machine_schedule[assigned_machine].append((earliest_start, completion_time, current_operation))
            
            # 更新工件完成时间
            job_completion_times[selected_job] = completion_time
            
            # 更新机器完成时间
            machine_completion_times[assigned_machine] = max(
                machine_completion_times[assigned_machine], completion_time
            )
            
            # 更新该工件的当前工序索引
            job_current_operation[selected_job] += 1
            scheduled_count += 1
        
        makespan = max(machine_completion_times.values()) if machine_completion_times else 0
        return makespan, start_times
    
    def _find_earliest_available_time(self, machine_schedule: List[Tuple[int, int, FlexibleOperation]], 
                                    processing_time: int, earliest_start_job: int) -> int:
        """
        在机器调度表中找到能够容纳指定处理时间的最早可用时间
        
        Args:
            machine_schedule: 机器的调度表 [(start_time, end_time, operation), ...]
            processing_time: 需要的处理时间
            earliest_start_job: 工件约束的最早开始时间
            
        Returns:
            int: 最早可用的开始时间
        """
        if not machine_schedule:
            return earliest_start_job
        
        # 创建副本避免修改原始数据，按开始时间排序
        schedule_copy = sorted(machine_schedule, key=lambda x: x[0])
        
        # 检查是否可以在第一个工序之前安排
        if schedule_copy[0][0] >= earliest_start_job + processing_time:
            return earliest_start_job
        
        # 检查相邻工序之间的空隙
        for i in range(len(schedule_copy) - 1):
            gap_start = max(schedule_copy[i][1], earliest_start_job)  # 考虑工件约束
            gap_end = schedule_copy[i + 1][0]  # 后一个工序的开始时间
            
            if gap_end - gap_start >= processing_time:
                return gap_start
        
        # 如果没有合适的空隙，安排在最后一个工序之后或工件约束时间之后
        return max(schedule_copy[-1][1], earliest_start_job)

    
    def _initialize_population(self, size: int) -> List[Dict[str, str]]:
        """初始化种群"""
        population = []
        
        for _ in range(size):
            # 随机机器分配
            machine_assignment = {}
            for op in self.schedule.operations:
                available_machines = op.get_available_machines()
                machine_assignment[op.id] = random.choice(available_machines)
            
            population.append(machine_assignment)
        
        return population
    
    def _roulette_selection(self, population: List, fitness_scores: List[float]):
        """轮盘赌选择"""
        total_fitness = sum(fitness_scores)
        if total_fitness == 0:
            return random.choice(population)
        
        pick = random.uniform(0, total_fitness)
        current = 0
        for i, fitness in enumerate(fitness_scores):
            current += fitness
            if current >= pick:
                return population[i]
        return population[-1]
    
    def _crossover(self, parent1: Dict[str, str], parent2: Dict[str, str]) -> Tuple[Dict[str, str], Dict[str, str]]:
        """交叉操作"""
        child_assignment1 = parent1.copy()
        child_assignment2 = parent2.copy()
        
        # 随机选择一部分工序进行交换
        operations = list(parent1.keys())
        crossover_point = len(operations) // 2
        random.shuffle(operations)
        
        for op_id in operations[:crossover_point]:
            child_assignment1[op_id] = parent2[op_id]
            child_assignment2[op_id] = parent1[op_id]
        
        return child_assignment1, child_assignment2
    
    def _mutate(self, individual: Dict[str, str]) -> Dict[str, str]:
        """变异操作"""
        new_assignment = individual.copy()
        
        # 随机改变几个工序的机器分配
        operations = list(individual.keys())
        num_mutations = max(1, len(operations) // 20)  # 减少变异强度
        
        for _ in range(num_mutations):
            op_id = random.choice(operations)
            # 找到对应的工序对象
            operation = None
            for op in self.schedule.operations:
                if op.id == op_id:
                    operation = op
                    break
            
            if operation:
                available_machines = operation.get_available_machines()
                if len(available_machines) > 1:  # 只有多个选择时才变异
                    current_machine = new_assignment[op_id]
                    available_machines = [m for m in available_machines if m != current_machine]
                    if available_machines:                        new_assignment[op_id] = random.choice(available_machines)
        
        return new_assignment
    
    def genetic_algorithm(self, population_size: int = 100, generations: int = 500,
                         mutation_rate: float = 0.1, crossover_rate: float = 0.8) -> Tuple[int, Dict, Dict]:
        """
        遗传算法求解FJSP - 不使用machine_orders
        """
        print(f"开始遗传算法求解 (种群大小: {population_size}, 代数: {generations})")
        
        # 初始化种群
        population = self._initialize_population(population_size)
        
        best_makespan = float('inf')
        best_solution = None
        best_start_times = None
        generation_bests = []
        
        for generation in range(generations):
            # 评估种群适应度
            fitness_scores = []
            for individual in population:
                machine_assignment = individual
                makespan, start_times = self.calculate_makespan(machine_assignment)
                
                # 计算负载平衡因子
                machine_loads = defaultdict(int)
                for op in self.schedule.operations:
                    assigned_machine = machine_assignment[op.id]
                    machine_loads[assigned_machine] += op.get_processing_time(assigned_machine)
                
                if machine_loads:
                    avg_load = sum(machine_loads.values()) / len(machine_loads)
                    load_variance = sum((load - avg_load)**2 for load in machine_loads.values()) / len(machine_loads)
                    balance_factor = 1 / (1 + load_variance / (avg_load**2 + 1))
                else:
                    balance_factor = 1
                
                # 综合适应度：考虑makespan和负载平衡
                fitness = (1.0 / (makespan + 1)) * (0.7 + 0.3 * balance_factor)
                fitness_scores.append(fitness)
                
                if makespan < best_makespan:
                    best_makespan = makespan
                    best_start_times = start_times
                    best_assignment = machine_assignment.copy()

            generation_bests.append(best_makespan)
            
            # 选择、交叉、变异
            new_population = []
            
            # 保留最优个体（精英策略）
            best_idx = fitness_scores.index(max(fitness_scores))
            new_population.append(population[best_idx])
            
            # 生成新个体
            while len(new_population) < population_size:
                # 轮盘赌选择
                parent1 = self._roulette_selection(population, fitness_scores)
                parent2 = self._roulette_selection(population, fitness_scores)
                
                # 交叉
                if random.random() < crossover_rate:
                    child1, child2 = self._crossover(parent1, parent2)
                else:
                    child1, child2 = parent1, parent2
                
                # 变异
                if random.random() < mutation_rate:
                    child1 = self._mutate(child1)
                if random.random() < mutation_rate:
                    child2 = self._mutate(child2)
                
                new_population.extend([child1, child2])
            
            population = new_population[:population_size]
            
            print(f"第 {generation} 代，最佳makespan: {best_makespan}")
        
        print(f"遗传算法完成，最佳makespan: {best_makespan}")
        
        # 绘制收敛曲线
        self._plot_convergence(generation_bests)
          # 计算machine_orders用于可视化
        machine_orders = self._calculate_machine_orders(best_assignment, best_start_times)
        
        return best_makespan, best_assignment, machine_orders, best_start_times
    
    def _plot_convergence(self, generation_bests: List[int]):
        """绘制算法收敛曲线"""
        plt.figure(figsize=(10, 6))
        plt.plot(generation_bests, 'b-', linewidth=2, alpha=0.8)
        plt.xlabel('Generation', fontsize=12)
        plt.ylabel('Best Makespan', fontsize=12)
        plt.title('遗传算法收敛曲线', fontsize=14)
        plt.grid(True, alpha=0.3)
        plt.tight_layout()
        plt.show()
    
    def _calculate_machine_orders(self, machine_assignment: Dict[str, str], 
                                 start_times: Dict[str, int]) -> Dict[str, List[FlexibleOperation]]:
        """基于机器分配和开始时间计算机器工序顺序"""
        machine_orders = defaultdict(list)
        
        # 为每台机器收集工序
        for op in self.schedule.operations:
            if op.id in machine_assignment and op.id in start_times:
                assigned_machine = machine_assignment[op.id]
                print(f"工序 {op.id} 分配到机器 {assigned_machine}，开始时间 {start_times[op.id]}")
                machine_orders[assigned_machine].append((start_times[op.id], op))
        
        # 按开始时间排序
        for machine in machine_orders:
            machine_orders[machine].sort(key=lambda x: x[0])
            machine_orders[machine] = [op for _, op in machine_orders[machine]]
        
        return dict(machine_orders)


def read_real_case_data(filename: str) -> Tuple[List[FlexibleJob], Dict[str, str]]:
    """
    读取实际案例Excel数据
    
    Args:
        filename: Excel文件路径
        
    Returns:
        (jobs, machine_mapping): 柔性工件列表和机器映射表
    """

    # 使用openpyxl读取Excel文件
    workbook = openpyxl.load_workbook(filename)
    worksheet = workbook.active
    
    print(f"读取Excel文件: {filename}")
    print(f"工作表名称: {worksheet.title}")
    print(f"数据范围: {worksheet.max_row} 行 x {worksheet.max_column} 列")
    
    # 解析数据结构
    jobs = []
    machine_mapping = {}
    
    # 从第5行开始读取数据
    current_job_id = 1
    current_job_operations = []
    current_quantity = 1
    
    # 机器列名映射 (F列开始对应所有机器，E列是工序序号)
    machine_columns = MACHINE_LIST
    part_name = ""
    for row in range(6, worksheet.max_row + 1):  # 从第6行开始

        # 读取各列数据
        cur_part_name = worksheet.cell(row=row, column=2).value  # B列：零件名称
        
        cur_quantity = worksheet.cell(row=row, column=3).value   # C列：数量
        feature = worksheet.cell(row=row, column=4).value    # D列：特征
        operation_seq = worksheet.cell(row=row, column=5).value  # E列：工序序号
        feature = str(feature).strip() if feature else ""
        
        # 处理数量
        quantity = int(float(str(cur_quantity).strip())) if cur_quantity else quantity
        
        # 处理工序序号
        operation_seq = int(float(str(operation_seq).strip())) if operation_seq else 0

        # 检查是否开始新工件（B列出现新值）
        if cur_part_name is not None and part_name != cur_part_name:
            
            # 创建前一个工件（根据数量创建多个相同工件）
            if current_job_operations:
                for i in range(current_quantity):
                    # 为每个工件创建独立的工序副本，使用不同的job_id
                    job_operations = []
                    for op in current_job_operations:
                        new_op = FlexibleOperation(
                            current_job_id, 
                            op.operation_id, 
                            op.machine_options.copy(), 
                            part_name
                        )
                        job_operations.append(new_op)
                    
                    job = FlexibleJob(current_job_id, job_operations, part_name)
                    jobs.append(job)
                    current_job_id += 1
            part_name = str(cur_part_name).strip() if cur_part_name else part_name
            current_job_operations = []

        current_quantity = quantity
        
        # 解析机器选项和时间（从F列开始）
        machine_options = []
        
        # 读取F列到后续列的机器时间数据
        for i, machine_name in enumerate(machine_columns):
            col_idx = 6 + i  # F列开始（索引6）
            cell_value = worksheet.cell(row=row, column=col_idx).value
            
            if cell_value is not None and str(cell_value).strip() != '':
                try:
                    time_val = int(float(str(cell_value).strip()))
                    if time_val > 0:  # 只添加正数时间
                        machine_options.append((machine_name, time_val))
                except (ValueError, TypeError):
                    continue

        # 如果有有效的机器选项，创建工序
        if machine_options and operation_seq > 0:
            # 使用临时的job_id创建模板工序
            operation = FlexibleOperation(0, operation_seq, machine_options, part_name)
            current_job_operations.append(operation)
    
    # 添加最后一个工件（根据数量创建多个相同工件）
    if current_job_operations:
        for i in range(current_quantity):
            # 为每个工件创建独立的工序副本，使用不同的job_id
            job_operations = []
            for op in current_job_operations:
                new_op = FlexibleOperation(
                    current_job_id, 
                    op.operation_id, 
                    op.machine_options.copy(), 
                    part_name
                )
                job_operations.append(new_op)
            
            job = FlexibleJob(current_job_id, job_operations, part_name)
            jobs.append(job)
            current_job_id += 1

    # 创建机器映射
    all_machines = set()
    for job in jobs:
        for op in job.operations:
            for machine, _ in op.machine_options:
                all_machines.add(machine)
    
    machine_mapping = {machine: machine for machine in sorted(all_machines)}
    
    print(f"\n解析完成:")
    print(f"工件数量: {len(jobs)}")
    print(f"机器数量: {len(machine_mapping)}")
    print(f"机器列表: {sorted(machine_mapping.keys())}")
    
    workbook.close()
    return jobs, machine_mapping


def visualize_flexible_schedule(machine_assignment: Dict[str, str],
                               machine_orders: Dict[str, List[FlexibleOperation]], 
                               start_times: Dict[str, int], 
                               title: str = "柔性作业车间调度甘特图"):
    """
    可视化柔性调度方案的甘特图 - 修复显示问题
    """
    if not machine_orders or not start_times:
        print("无效的调度数据，无法生成甘特图")
        return
        
    fig, ax = plt.subplots(figsize=(20, 12))
    
    # 扩展颜色映射
    colors = [
        '#FF6B6B', '#4ECDC4', '#45B7D1', '#96CEB4', '#FFEAA7', 
        '#DDA0DD', '#98D8C8', '#F7DC6F', '#BB8FCE', '#85C1E9',
        '#F8C471', '#82E0AA', '#F1948A', '#85C1E9', '#D2B4DE',
        '#AED6F1', '#A3E4D7', '#D5DBDB', '#FADBD8', '#D1F2EB',
        '#FCF3CF', '#EBDEF0', '#EAF2F8', '#E8F8F5', '#FDF2E9'
    ]
    
    # 为不同产品分配颜色
    product_colors = {}
    job_colors = {}
    
    machines = sorted(machine_orders.keys())
    y_positions = {machine: i for i, machine in enumerate(machines)}
    
    # 计算最大时间
    max_time = 0
    for ops in machine_orders.values():
        for op in ops:
            if op.id in start_times:
                finish_time = start_times[op.id] + op.get_processing_time(machine_assignment[op.id])
                max_time = max(max_time, finish_time)
    
    print(f"甘特图信息:")
    print(f"  机器数量: {len(machines)}")
    print(f"  总工序数: {sum(len(ops) for ops in machine_orders.values())}")
    print(f"  最大时间: {max_time}")
    
    # 绘制甘特图
    for machine, operations in machine_orders.items():
        y_pos = y_positions[machine]
        print(f"  机器 {machine}: {len(operations)} 个工序")
        
        for i, op in enumerate(operations):
            # 为产品类型分配颜色
            if op.product_type not in product_colors:
                product_colors[op.product_type] = colors[len(product_colors) % len(colors)]
            
            # 为工件分配相同产品的颜色
            if op.job_id not in job_colors:
                base_color = product_colors[op.product_type]
                job_colors[op.job_id] = base_color
            
            if op.id in start_times:
                start_time = start_times[op.id]
                duration = op.get_processing_time(machine_assignment[op.id])
                
                # print(f"工序 {op.id}: 开始={start_time}, 持续={duration}, 结束={start_time + duration}, 机器={machine_assignment[op.id]}")
                
                # 绘制工序矩形
                rect = patches.Rectangle((start_time, y_pos - 0.35), duration, 0.7,
                                       linewidth=1.5, edgecolor='white', 
                                       facecolor=job_colors[op.job_id], alpha=0.8)
                ax.add_patch(rect)
                
                # 添加工序标签
                if duration > max_time * 0.01:  # 降低显示文字的阈值
                    ax.text(start_time + duration/2, y_pos, f'J{op.job_id}O{op.operation_id}',
                           ha='center', va='center', fontsize=9, fontweight='bold', color='black')
    
    # 设置图表属性
    ax.set_xlabel('Time (时间单位)', fontsize=16, fontweight='bold')
    ax.set_ylabel('Machine (机器)', fontsize=16, fontweight='bold')
    ax.set_title(title, fontsize=18, fontweight='bold', pad=20)
    
    # 设置y轴
    ax.set_yticks(range(len(machines)))
    ax.set_yticklabels([f'{m}' for m in machines], fontsize=12)
    ax.set_ylim(-0.5, len(machines) - 0.5)
    
    # 设置x轴
    if max_time > 0:
        ax.set_xlim(0, max_time + max_time * 0.05)
        # 添加时间刻度
        time_ticks = range(0, int(max_time) + 1, max(1, int(max_time) // 20))
        ax.set_xticks(time_ticks)
    ax.tick_params(axis='x', labelsize=12)
    ax.tick_params(axis='y', labelsize=12)
    
    # 添加网格
    ax.grid(True, alpha=0.3, linestyle='-', linewidth=0.5)
    ax.set_axisbelow(True)
    
    # 添加产品图例
    legend_elements = []
    for product_type in sorted(product_colors.keys()):
        legend_elements.append(
            patches.Patch(facecolor=product_colors[product_type], 
                         label=f'{product_type}', alpha=0.8,
                         edgecolor='white', linewidth=1)
        )
    
    if legend_elements:
        ax.legend(handles=legend_elements, loc='center left', 
                 bbox_to_anchor=(1.02, 0.5), fontsize=12,
                 title='产品类型', title_fontsize=14)
    
    # 添加makespan标记线
    if max_time > 0:
        ax.axvline(x=max_time, color='red', linestyle='--', linewidth=3, alpha=0.8)
        ax.text(max_time + max_time * 0.01, len(machines) * 0.95, 
               f'Makespan = {max_time}', 
               fontsize=14, fontweight='bold', color='red',
               bbox=dict(boxstyle="round,pad=0.3", facecolor="white", alpha=0.8))
    
    # 添加机器利用率信息
    machine_utilization = {}
    for machine in machines:
        total_work_time = 0
        if machine in machine_orders:
            for op in machine_orders[machine]:
                if op.id in start_times:
                    total_work_time += op.get_processing_time(machine_assignment[op.id])
        utilization = (total_work_time / max_time * 100) if max_time > 0 else 0
        machine_utilization[machine] = utilization
    
    # 在图表上显示利用率
    for i, machine in enumerate(machines):
        util = machine_utilization[machine]
        ax.text(-max_time * 0.02, i, f'{util:.1f}%', 
               ha='right', va='center', fontsize=10, 
               bbox=dict(boxstyle="round,pad=0.2", facecolor="lightblue", alpha=0.7))
    
    ax.set_facecolor('#FAFAFA')
    plt.tight_layout()
    plt.show()
    
    # 打印详细的调度统计
    print(f"\n机器利用率统计:")
    for machine in sorted(machines):
        util = machine_utilization[machine]
        work_time = sum(op.get_processing_time(machine_assignment[op.id]) 
                       for op in machine_orders.get(machine, []) 
                       if op.id in start_times)
        print(f"  {machine}: 工作时间={work_time}, 利用率={util:.1f}%")

def save_solution_to_file(machine_assignment: Dict[str, str],
                         machine_orders: Dict[str, List[FlexibleOperation]], 
                         start_times: Dict[str, int],
                         flexible_jobs: List[FlexibleJob],
                         filename: str = "realcase_sol.txt"):
    """
    将调度方案保存到文件中，按指定格式输出
    
    Args:
        machine_assignment: 机器分配方案
        machine_orders: 机器工序顺序
        start_times: 工序开始时间
        flexible_jobs: 柔性工件列表
        filename: 输出文件名
    """
    print(f"\n正在保存调度方案到文件: {filename}")
    
    # 创建解决方案记录列表
    solution_records = []
    
    # 创建产品类型到零件编号的映射
    product_to_number = {}
    unique_products = set()
    for job in flexible_jobs:
        unique_products.add(job.product_type)
    
    # 按产品类型排序并分配编号
    for i, product_type in enumerate(sorted(unique_products), 1):
        product_to_number[product_type] = i
    
    # 计算工件数量和机器数量
    num_jobs = len(flexible_jobs)
    num_machines = len(set(machine_assignment.values()))
    
    # 遍历每台机器的工序安排
    for machine_name, operations in machine_orders.items():
        for op in operations:
            if op.id in start_times:
                start_time = start_times[op.id]
                processing_time = op.get_processing_time(machine_assignment[op.id])
                end_time = start_time + processing_time
                
                # 获取零件编号（基于产品类型）
                part_number = op.job_id
                
                solution_records.append({
                    'equipment': machine_name,
                    'part_number': part_number,
                    'operation_number': op.operation_id,
                    'start_time': start_time,
                    'end_time': end_time,
                    'product_type': op.product_type,
                    'job_id': op.job_id
                })
    
    # 按设备编号和开始时间排序
    solution_records.sort(key=lambda x: (x['equipment'], x['start_time']))
    
    # 写入文件
    with open(filename, 'w', encoding='utf-8') as f:
        # 写入工件数量和机器数量信息
        f.write(f"Nb of jobs {num_jobs} Nb of Machines {num_machines}\n")
        
        # 写入表头
        f.write("Solution\n")
        f.write(f"{'设备编号':<12} {'零件编号':<12} {'工序标号':<12} {'开始时间':<12} {'结束时间':<12}\n")
        
        # 写入调度记录
        current_equipment = None
        for record in solution_records:
            equipment = record['equipment']
            part_num = record['part_number']
            operation_num = record['operation_number']
            start_time = record['start_time']
            end_time = record['end_time']

            f.write(f"{equipment:<12} {part_num:<12} {operation_num:<12} {start_time:<12} {end_time:<12}\n")
        


def solve_real_case(max_iterations: int = 5000):
    """求解实际案例"""
    print("开始处理实际案例数据...")
    
    # 读取Excel数据
    filename = "data/real_case.xlsx"
    flexible_jobs, machine_mapping = read_real_case_data(filename)
    
    if not flexible_jobs:
        print("未能成功读取数据，请检查Excel文件格式")
        return
    
    graph = FlexibleSchedule(flexible_jobs)
    solver = FlexibleScheduleSolver(graph)
    
    print(f"工序总数: {len(graph.operations)}")
    print(f"机器总数: {len(graph.all_machines)}")
    print(f"机器列表: {sorted(graph.all_machines)}")

    
    # 显示产品类型统计
    product_stats = defaultdict(int)
    for job in flexible_jobs:
        product_stats[job.product_type] += 1
    
    print(f"\n产品类型统计:")
    for product, count in sorted(product_stats.items()):
        print(f"  {product}: {count} 个工件")
    
    # 运行遗传算法
    print(f"\n开始求解柔性作业车间调度问题...")
    start_time = time.time()

    best_makespan, best_assignment, best_orders, best_start_times = solver.genetic_algorithm(
        population_size=100,
        generations=max_iterations,
        mutation_rate=0.12,
        crossover_rate=0.88
    )
    
    solve_time = time.time() - start_time
    
    if best_makespan != float('inf'):
        print(f"\n求解完成!")
        print(f"最优makespan: {best_makespan}")
        print(f"求解时间: {solve_time:.3f}秒")

        
        # 保存调度方案到文件
        save_solution_to_file(best_assignment, best_orders, best_start_times, flexible_jobs, "realcase_sol.txt")
        
        # 生成甘特图
        visualize_flexible_schedule(
            best_assignment, best_orders, best_start_times,
            f"柔性作业车间调度结果 (Makespan = {best_makespan})"
        )


if __name__ == "__main__":
    # 设置中文字体支持
    plt.rcParams['font.sans-serif'] = ['SimHei', 'DejaVu Sans']
    plt.rcParams['axes.unicode_minus'] = False

    # 运行柔性作业车间调度实际案例
    solve_real_case(max_iterations=100)
