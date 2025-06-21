import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as patches
from collections import defaultdict, deque
import itertools
import random
import copy
from typing import List, Dict, Tuple, Set, Optional
import time

class Operation:
    """工序类 - 表示单个加工工序"""
    def __init__(self, job_id: int, op_index: int, machine: int, processing_time: int):
        self.job_id = job_id          # 所属工件ID
        self.op_index = op_index      # 在工件中的工序索引
        self.machine = machine        # 加工机器
        self.processing_time = processing_time  # 加工时间
        self.id = f"O_{job_id}_{op_index}"      # 工序唯一标识
    
    def __str__(self):
        return f"{self.id}(J{self.job_id}, M{self.machine}, t={self.processing_time})"
    
    def __repr__(self):
        return self.__str__()

class Job:
    """工件类 - 表示完整的工件及其工艺路径"""
    def __init__(self, job_id: int, machines: List[int], processing_times: List[int]):
        self.job_id = job_id
        self.operations = []
        
        # 创建工序序列
        for i, (machine, time) in enumerate(zip(machines, processing_times)):
            op = Operation(job_id, i+1, machine, time)
            self.operations.append(op)
    
    def get_operation(self, op_index: int) -> Operation:
        """获取指定索引的工序"""
        return self.operations[op_index - 1]
    
    def __str__(self):
        return f"Job {self.job_id}: {' -> '.join([str(op) for op in self.operations])}"

class DisjunctiveGraph:
    """析取图类 - 核心数据结构"""
    def __init__(self, jobs: List[Job]):
        self.jobs = jobs
        self.operations = []
        self.conjunctive_arcs = []  # 连接弧（固定工艺顺序）
        self.disjunctive_arcs = []  # 析取弧（机器资源竞争）
        self.machine_operations = defaultdict(list)  # 每台机器上的工序
        
        self._build_graph()
    
    def _build_graph(self):
        """构建析取图的节点和边"""
        # 收集所有工序
        for job in self.jobs:
            self.operations.extend(job.operations)
            
        # 按机器分组工序
        for op in self.operations:
            self.machine_operations[op.machine].append(op)
        
        # 构建连接弧（工艺路径约束）
        self._build_conjunctive_arcs()
        
        # 构建析取弧（机器资源约束）
        self._build_disjunctive_arcs()
    
    def _build_conjunctive_arcs(self):
        """构建连接弧 - 表示工件内部工序的固定顺序"""
        for job in self.jobs:
            for i in range(len(job.operations) - 1):
                from_op = job.operations[i]
                to_op = job.operations[i + 1]
                self.conjunctive_arcs.append((from_op, to_op, from_op.processing_time))
    
    def _build_disjunctive_arcs(self):
        """构建析取弧 - 表示同一机器上工序间的竞争关系"""
        for machine, ops in self.machine_operations.items():
            # 对每台机器上的工序两两组合创建析取弧
            for i in range(len(ops)):
                for j in range(i + 1, len(ops)):
                    op1, op2 = ops[i], ops[j]
                    # 创建双向析取弧
                    self.disjunctive_arcs.append((op1, op2, op1.processing_time))
                    self.disjunctive_arcs.append((op2, op1, op2.processing_time))
    
    def is_acyclic(self, machine_orders: Dict[int, List[Operation]]) -> bool:
         # 构建有向图
        graph, in_degree = self._build_directed_graph(machine_orders)
        
        # 执行拓扑排序检测环
        is_acyclic_result, _ = self._topological_sort(graph, in_degree, compute_distances=False)
        
        return is_acyclic_result
    
    def sample_valid_machine_order(self, max_attempts: int = 1000) -> Dict[int, List[Operation]]:
        """
        高效采样有效的机器工序顺序 - 使用启发式方法提高效率
        
        Args:
            max_attempts: 最大尝试次数
            
        Returns:
            Dict[int, List[Operation]]: 有效的机器工序顺序，失败返回False
        """
        for attempt in range(max_attempts):
            machine_orders = {}
            
            # 随机打乱机器处理顺序
            machines = list(self.machine_operations.keys())
            random.shuffle(machines)
            
            # 为每台机器生成顺序
            valid = True
            for machine in machines:
                ops = self.machine_operations[machine]
                if len(ops) <= 1:
                    machine_orders[machine] = ops.copy()
                    continue
                
                # 使用启发式方法生成初始顺序
                order = self._generate_heuristic_order(ops, machine_orders)
                
                # 如果启发式方法失败，使用随机方法
                if not order:
                    order = ops.copy()
                    random.shuffle(order)
                
                machine_orders[machine] = order
                
                # 增量检查是否有环（提前终止）
                if not self.is_acyclic(machine_orders):
                    valid = False
                    break
            
            if valid and self.is_acyclic(machine_orders):
                return machine_orders
        
        return False
    
    def _generate_heuristic_order(self, operations: List[Operation], 
                                current_orders: Dict[int, List[Operation]]) -> List[Operation]:
        """
        使用启发式方法生成机器上的工序顺序
        
        Args:
            operations: 待排序的工序列表
            current_orders: 当前已确定的机器顺序
            
        Returns:
            List[Operation]: 启发式排序结果
        """
        if len(operations) <= 1:
            return operations.copy()
        
        # 计算每个工序的优先级分数
        scores = []
        for op in operations:
            score = 0
            
            # 基于工序在工件中的位置（越早的工序优先级越高）
            score += (len(self.jobs[op.job_id - 1].operations) - op.op_index) * 10
            
            # 基于加工时间（短加工时间优先 - SPT规则）
            score += (100 - op.processing_time)
            
            # 基于关键路径估计
            score += self._estimate_critical_path_priority(op, current_orders)
            
            # 添加随机扰动避免过度确定性
            score += random.uniform(-5, 5)
            
            scores.append((score, op))
        
        # 按分数排序
        scores.sort(key=lambda x: x[0], reverse=True)
        return [op for _, op in scores]
    
    def _estimate_critical_path_priority(self, operation: Operation, 
                                       current_orders: Dict[int, List[Operation]]) -> float:
        """
        估计工序在关键路径中的优先级
        
        Args:
            operation: 目标工序
            current_orders: 当前机器顺序
            
        Returns:
            float: 优先级分数
        """
        # 计算该工序所属工件的总加工时间
        job = self.jobs[operation.job_id - 1]
        total_job_time = sum(op.processing_time for op in job.operations)
        
        # 计算工序前置时间（该工序之前的所有工序加工时间）
        preceding_time = sum(op.processing_time for op in job.operations[:operation.op_index-1])
        
        # 计算工序后续时间
        following_time = sum(op.processing_time for op in job.operations[operation.op_index:])
        
        # 关键路径估计：前置时间少且后续时间多的工序优先级高
        priority = following_time - preceding_time * 0.5
        
        return priority
    
    
    def _build_directed_graph(self, machine_orders: Dict[int, List[Operation]]) -> Tuple[defaultdict, defaultdict]:
        """
        构建有向图的邻接表和入度信息
        
        Args:
            machine_orders: 每台机器上工序的加工顺序
            
        Returns:
            (graph, in_degree): 邻接表和入度字典
        """
        graph = defaultdict(list)
        in_degree = defaultdict(int) 
        
        # 初始化所有工序的入度（包括虚拟节点）
        for op in self.operations:
            in_degree[op.id] = 0
        
        # in_degree[self.start_node.id] = 0

        # 添加工艺路径约束边（连接弧，包括虚拟弧）
        for from_op, to_op, weight in self.conjunctive_arcs:
            graph[from_op.id].append((to_op.id, weight))
            in_degree[to_op.id] += 1
        
        # 添加机器顺序约束边（选择的析取弧）
        for machine, ops in machine_orders.items():
            for i in range(len(ops) - 1):
                from_op = ops[i]
                to_op = ops[i + 1]
                graph[from_op.id].append((to_op.id, from_op.processing_time))
                in_degree[to_op.id] += 1
        # print(in_degree)
        # print(graph)
        return graph, in_degree
    
    def _topological_sort(self, graph: defaultdict, in_degree: defaultdict, 
                         compute_distances: bool = False) -> Tuple[bool, Optional[Dict[str, int]]]:
        """
        执行拓扑排序
        
        Args:
            graph: 邻接表，格式为 {node_id: [(neighbor_id, weight), ...]}
            in_degree: 入度字典
            compute_distances: 是否计算最长距离
            
        Returns:
            (is_acyclic, distances): 是否无环，以及距离字典（如果需要计算）
        """
        queue = deque()
        distances = {} if compute_distances else None
    
        
        # 找到所有入度为0的节点
        for op_id in in_degree:
            if in_degree[op_id] == 0:
                queue.append(op_id)
        # print(f"初始入度为0的节点: {list(queue)}")
        processed_count = 0
        
        while queue:
            current_id = queue.popleft()
            processed_count += 1
            
            # 处理当前节点的所有后继
            for neighbor_info in graph[current_id]:
                # print(f"Processing edge {current_id} -> {neighbor_info}")
                if isinstance(neighbor_info, tuple):
                    neighbor_id, weight = neighbor_info
                else:
                    neighbor_id, weight = neighbor_info, 0
                
                in_degree[neighbor_id] -= 1
                if in_degree[neighbor_id] == 0:
                    queue.append(neighbor_id)
        
        # 检查是否无环
        # print(f"Processed {processed_count} nodes out of {len(self.operations)}")
        is_acyclic = processed_count == len(self.operations)
        
        return is_acyclic, distances

    
class ScheduleSolver:
    """调度求解器 - 实现多种优化算法"""
    def __init__(self, disjunctive_graph: DisjunctiveGraph):
        self.graph = disjunctive_graph
        self.best_makespan = float('inf')
        self.best_schedule = None
    
    def calculate_makespan(self, machine_orders: Dict[int, List[Operation]]) -> Tuple[int, Dict[str, int]]:
        """
        计算给定机器顺序下的最大完工时间（仅处理无环图）
        
        Args:
            machine_orders: 每台机器上工序的加工顺序
            
        Returns:
            (makespan, operation_start_times): 最大完工时间和各工序开始时间
        """
        if not machine_orders:
            return float('inf'), {}
        
        # 检查是否为无环图，有环直接返回无穷大
        if not self.graph.is_acyclic(machine_orders):
            return float('inf'), {}
        
        # 构建有向图
        graph, in_degree = self.graph._build_directed_graph(machine_orders)
        
        # 使用拓扑排序计算最早开始时间
        queue = deque()
        start_times = {}
        
        # 初始化所有工序的开始时间
        for op in self.graph.operations:
            start_times[op.id] = 0
        
        # 找到所有入度为0的节点
        for op_id in in_degree:
            if in_degree[op_id] == 0:
                queue.append(op_id)
        
        while queue:
            current_id = queue.popleft()
            
            # 处理当前节点的所有后继
            for next_id, processing_time in graph[current_id]:
                # 更新后继节点的最早开始时间
                start_times[next_id] = max(start_times[next_id], 
                                         start_times[current_id] + processing_time)
                
                in_degree[next_id] -= 1
                if in_degree[next_id] == 0:
                    queue.append(next_id)
        
        # 计算最大完工时间
        makespan = 0
        for op in self.graph.operations:
            finish_time = start_times[op.id] + op.processing_time
            makespan = max(makespan, finish_time)
        
        return makespan, start_times
    
    
    def random_search(self, max_iterations: int = 1000, max_attempts=2000) -> Tuple[int, Dict[int, List[Operation]]]:
        """
        随机搜索算法 - 替代遗传算法用于大规模问题
        
        Args:
            max_iterations: 最大迭代次数
            max_attempts: 每次采样的最大尝试次数
            
        Returns:
            (best_makespan, best_machine_orders): 最优解和对应的机器顺序
        """
        print(f"开始随机搜索求解 (最大迭代次数: {max_iterations})...")
        
        best_makespan = float('inf')
        best_orders = None
        valid_solutions = 0
        
        for iteration in range(max_iterations):
            # 生成随机调度方案
            machine_orders = self.graph.sample_valid_machine_order(max_attempts)
            if not machine_orders:
                continue
                
            # 计算makespan
            makespan, _ = self.calculate_makespan(machine_orders)
            
            if makespan != float('inf'):
                valid_solutions += 1
                
                if makespan < best_makespan:
                    best_makespan = makespan
                    best_orders = machine_orders.copy()
                    print(f"第 {iteration+1} 次迭代发现更优解: Makespan = {best_makespan}")

            # 每1000次迭代输出进度
            if (iteration + 1) % 1000 == 0:
                print(f"已完成 {iteration+1}/{max_iterations} 次迭代, 有效解: {valid_solutions}, 当前最优解: {best_makespan}")
        
        print(f"随机搜索完成! 总有效解: {valid_solutions}/{max_iterations}, 最优 Makespan: {best_makespan}")
        return best_makespan, best_orders

def read_benchmark_data(filename: str) -> Tuple[List[Tuple[List[int], List[int]]], int, int]:
    """
    读取标准测试数据文件 - 改进版本，处理机器编号问题
    
    Args:
        filename: 数据文件路径
        
    Returns:
        (job_data, num_jobs, num_machines): 工件数据、工件数量、机器数量
    """
    with open(filename, 'r') as file:
        lines = file.readlines()
    
    # 读取第一行获取工件数和机器数
    first_line = lines[0].strip()
    parts = first_line.split()
    num_jobs = int(parts[3])
    num_machines = int(parts[7])
    
    print(f"读取数据文件: {filename}")
    print(f"工件数量: {num_jobs}, 机器数量: {num_machines}")
    
    # 找到Times和Machines部分
    times_start = -1
    machines_start = -1
    
    for i, line in enumerate(lines):
        if line.strip() == "Times":
            times_start = i + 1
        elif line.strip() == "Machines":
            machines_start = i + 1
            break
    
    if times_start == -1 or machines_start == -1:
        raise ValueError("文件格式错误：找不到Times或Machines部分")
    
    # 读取加工时间数据
    times_data = []
    for i in range(times_start, machines_start - 1):
        line = lines[i].strip()
        if line:
            times = [int(x) for x in line.split()]
            times_data.extend(times)
    
    # 读取机器分配数据
    machines_data = []
    for i in range(machines_start, len(lines)):
        line = lines[i].strip()
        if line:
            machines = [int(x) for x in line.split()]
            machines_data.extend(machines)
    
    # 验证数据长度
    expected_length = num_jobs * num_machines
    if len(times_data) != expected_length:
        print(f"警告: 时间数据长度 {len(times_data)} 不等于期望长度 {expected_length}")
    if len(machines_data) != expected_length:
        print(f"警告: 机器数据长度 {len(machines_data)} 不等于期望长度 {expected_length}")
    
    # 检查并修正机器编号（确保从1开始）
    if machines_data:
        min_machine = min(machines_data)
        max_machine = max(machines_data)
        print(f"原始机器编号范围: [{min_machine}, {max_machine}]")
        
        # 如果机器编号从0开始，转换为从1开始
        if min_machine == 0:
            machines_data = [m + 1 for m in machines_data]
            print("已将机器编号从0-based转换为1-based")
        
        # 验证机器编号范围
        unique_machines = set(machines_data)
        expected_machines = set(range(1, num_machines + 1))
        if unique_machines != expected_machines:
            print(f"警告: 机器编号不连续")
            print(f"实际机器: {sorted(unique_machines)}")
            print(f"期望机器: {sorted(expected_machines)}")
            
            # 重新映射机器编号
            machine_mapping = {old: new for new, old in enumerate(sorted(unique_machines), 1)}
            machines_data = [machine_mapping.get(m, m) for m in machines_data]
            print(f"已重新映射机器编号: {machine_mapping}")
    
    # 重组数据为每个工件的格式
    job_data = []
    for job_id in range(num_jobs):
        job_machines = []
        job_times = []
        
        start_idx = job_id * num_machines
        end_idx = start_idx + num_machines
        
        # 确保索引不超出范围
        if end_idx <= len(machines_data) and end_idx <= len(times_data):
            job_machines = machines_data[start_idx:end_idx]
            job_times = times_data[start_idx:end_idx]
            
            # 验证工件数据
            if len(job_machines) != num_machines or len(job_times) != num_machines:
                print(f"警告: 工件 {job_id + 1} 数据不完整")
                continue
                
            # 检查是否有重复机器（在同一工件中）
            if len(set(job_machines)) != len(job_machines):
                print(f"警告: 工件 {job_id + 1} 包含重复的机器编号")
                
        else:
            print(f"错误: 工件 {job_id + 1} 的数据索引超出范围")
            continue
        
        job_data.append((job_machines, job_times))
    
    # 打印前几个工件的数据进行验证
    print(f"\n前3个工件的数据预览:")
    for i, (machines, times) in enumerate(job_data[:3]):
        print(f"工件 {i+1}: 机器序列 {machines}, 时间序列 {times}")
    
    return job_data, num_jobs, num_machines

def validate_job_data(job_data: List[Tuple[List[int], List[int]]], num_jobs: int, num_machines: int) -> List[Job]:
    """
    验证工件数据并创建工件对象
    
    Args:
        job_data: 工件数据列表
        num_jobs: 工件数量
        num_machines: 机器数量
        
    Returns:
        List[Job]: 验证后的工件对象列表
    """
    print("正在验证数据完整性...")
    
    # 检查数据合理性
    for i, (machines, times) in enumerate(job_data):
        if len(machines) != len(times):
            raise ValueError(f"工件 {i+1} 的机器和时间数据长度不匹配")
        
        if any(t <= 0 for t in times):
            print(f"警告: 工件 {i+1} 包含非正数加工时间")
        
        if any(m < 1 for m in machines):
            print(f"警告: 工件 {i+1} 包含无效机器编号")
    
    # 创建工件对象
    jobs = []
    for i, (machines, times) in enumerate(job_data):
        try:
            job = Job(i + 1, machines, times)
            jobs.append(job)
            if i < 3:
                print(f"工件 {i+1}: 机器序列长度={len(machines)}, 加工时间范围=[{min(times)}, {max(times)}]")
        except Exception as e:
            raise ValueError(f"创建工件 {i+1} 时发生错误: {e}")
    
    return jobs

def build_disjunctive_graph_and_solver(jobs: List[Job]) -> Tuple[DisjunctiveGraph, ScheduleSolver]:
    """
    构建析取图和求解器
    
    Args:
        jobs: 工件列表
        
    Returns:
        Tuple[DisjunctiveGraph, ScheduleSolver]: 析取图和求解器
    """
    try:
        graph = DisjunctiveGraph(jobs)
        solver = ScheduleSolver(graph)
        
        print(f"\n析取图构建完成:")
        print(f"- 工序总数: {len(graph.operations)}")
        print(f"- 连接弧数量: {len(graph.conjunctive_arcs)}")
        print(f"- 析取弧数量: {len(graph.disjunctive_arcs)}")
        
        # 验证机器操作分布
        print(f"- 各机器上的工序数量:")
        for machine in sorted(graph.machine_operations.keys()):
            ops_count = len(graph.machine_operations[machine])
            print(f"  机器 M{machine}: {ops_count} 个工序")
            
        return graph, solver
        
    except Exception as e:
        raise ValueError(f"构建析取图时发生错误: {e}")

def solve_and_visualize(solver: ScheduleSolver, dataset_name: str, num_jobs: int, max_iterations: int, max_attempts: int) -> Dict[str, Tuple]:
    """
    求解调度问题并可视化结果
    
    Args:
        solver: 调度求解器
        dataset_name: 数据集名称
        num_jobs: 工件数量
        max_iterations: 最大迭代次数
        max_attempts: 最大尝试次数
        
    Returns:
        Dict[str, Tuple]: 求解结果
    """
    print(f"\n运行随机搜索算法 (最大迭代次数: {max_iterations})...")
    
    try:
        start_time = time.time()
        best_makespan_rs, best_orders_rs = solver.random_search(max_iterations=max_iterations, max_attempts=max_attempts)
        rs_time = time.time() - start_time
        
        if best_makespan_rs != float('inf'):
            print(f"随机搜索完成: Makespan = {best_makespan_rs}, 时间 = {rs_time:.3f}s")
            
            # 生成甘特图
            print(f"\n正在为 {dataset_name} 生成甘特图...")
            try:
                _, start_times = solver.calculate_makespan(best_orders_rs)
                visualize_schedule(best_orders_rs, start_times, 
                                    f"{dataset_name} - 随机搜索 (Makespan = {best_makespan_rs})")
                
                # 显示调度详情
                if num_jobs <= 10:  # 只对小规模问题显示详细信息
                    print_schedule_details(best_orders_rs, start_times, best_makespan_rs)
                    
            except Exception as e:
                print(f"生成甘特图时出错: {e}")
                
            return {'随机搜索': (best_makespan_rs, rs_time)}
        else:
            print(f"随机搜索未找到有效解")
            return {'随机搜索': (None, None)}
            
    except Exception as e:
        print(f"随机搜索执行出错: {e}")
        return {'随机搜索': (None, None)}

def compare_algorithms_on_benchmark(job_data: List[Tuple[List[int], List[int]]], 
                                  dataset_name: str,
                                  num_jobs: int, 
                                  num_machines: int,
                                  max_iterations: int,
                                  max_attempts: int):
    """
    在标准数据集上比较不同算法的性能
    """
    print(f"\n{'='*80}")
    print(f"标准数据集 {dataset_name} 算法性能比较")
    print(f"规模: {num_jobs} 工件 × {num_machines} 机器")
    print(f"{'='*80}")
    
    try:
        # 验证数据并创建工件对象
        jobs = validate_job_data(job_data, num_jobs, num_machines)
        
        # 构建析取图和求解器
        graph, solver = build_disjunctive_graph_and_solver(jobs)
        
        # 求解并可视化
        results = solve_and_visualize(solver, dataset_name, num_jobs, max_iterations, max_attempts)
        
        return results
        
    except Exception as e:
        print(f"处理数据集 {dataset_name} 时发生错误: {e}")
        return None

def visualize_schedule(machine_orders: Dict[int, List[Operation]], 
                      start_times: Dict[str, int], 
                      title: str = "作业车间调度甘特图"):
    """
    可视化调度方案的甘特图
    
    Args:
        machine_orders: 每台机器的工序顺序
        start_times: 各工序的开始时间
        title: 图表标题
    """
    # 检查输入数据
    if not machine_orders or not start_times:
        print("无效的调度数据，无法生成甘特图")
        return
        
    fig, ax = plt.subplots(figsize=(14, 8))
    
    # 扩展颜色映射（为不同工件分配不同颜色）
    colors = [
        '#FF6B6B', '#4ECDC4', '#45B7D1', '#96CEB4', '#FFEAA7', 
        '#DDA0DD', '#98D8C8', '#F7DC6F', '#BB8FCE', '#85C1E9',
        '#F8C471', '#82E0AA', '#F1948A', '#85C1E9', '#D2B4DE',
        '#AED6F1', '#A3E4D7', '#D5DBDB', '#FADBD8', '#D1F2EB',
        '#FCF3CF', '#EBDEF0', '#EAF2F8', '#E8F8F5', '#FDF2E9'
    ]
    job_colors = {}
    
    machines = sorted(machine_orders.keys())
    y_positions = {machine: i for i, machine in enumerate(machines)}
    
    # 计算最大时间用于设置x轴
    max_time = max(start_times[op.id] + op.processing_time 
                  for ops in machine_orders.values() for op in ops)
    
    # 绘制甘特图
    for machine, operations in machine_orders.items():
        y_pos = y_positions[machine]
        
        for op in operations:
            job_id = op.job_id
            if job_id not in job_colors:
                job_colors[job_id] = colors[len(job_colors) % len(colors)]
            
            start_time = start_times[op.id]
            duration = op.processing_time
            
            # 绘制工序矩形 - 去掉文字标签
            rect = patches.Rectangle((start_time, y_pos - 0.35), duration, 0.7,
                                   linewidth=1.5, edgecolor='white', 
                                   facecolor=job_colors[job_id], alpha=0.8)
            ax.add_patch(rect)
    
    # 设置图表属性
    ax.set_xlabel('Time', fontsize=16, fontweight='bold')
    ax.set_ylabel('Machine', fontsize=16, fontweight='bold')
    ax.set_title(title, fontsize=18, fontweight='bold', pad=20)
    
    # 设置y轴
    ax.set_yticks(range(len(machines)))
    ax.set_yticklabels([f'M{m}' for m in machines], fontsize=14)
    ax.set_ylim(-0.5, len(machines) - 0.5)
    
    # 设置x轴
    ax.set_xlim(0, max_time + max_time * 0.05)  # 添加5%的边距
    ax.tick_params(axis='x', labelsize=12)
    ax.tick_params(axis='y', labelsize=12)
    
    # 添加网格
    ax.grid(True, alpha=0.3, linestyle='-', linewidth=0.5)
    ax.set_axisbelow(True)
    
    # 添加图例 - 显示工件信息
    legend_elements = []
    for job_id in sorted(job_colors.keys()):
        legend_elements.append(
            patches.Patch(facecolor=job_colors[job_id], 
                         label=f'Job {job_id}', alpha=0.8,
                         edgecolor='white', linewidth=1)
        )
    
    # 将图例放在图表右侧
    ax.legend(handles=legend_elements, loc='center left', 
             bbox_to_anchor=(1.02, 0.5), fontsize=12,
             title='Jobs', title_fontsize=14)
    
    # 添加makespan标记线
    ax.axvline(x=max_time, color='red', linestyle='--', linewidth=3, alpha=0.8)
    ax.text(max_time + max_time * 0.01, len(machines) * 0.95, 
           f'Makespan = {max_time}', 
           fontsize=14, fontweight='bold', color='red',
           bbox=dict(boxstyle="round,pad=0.3", facecolor="white", alpha=0.8))
    
    # 设置背景色
    ax.set_facecolor('#FAFAFA')
    
    plt.tight_layout()
    plt.show()

def print_schedule_details(machine_orders: Dict[int, List[Operation]], 
                          start_times: Dict[str, int], 
                          makespan: int):
    """
    打印调度方案的详细信息
    
    Args:
        machine_orders: 每台机器的工序顺序
        start_times: 各工序的开始时间
        makespan: 最大完工时间
    """
    print("\n" + "="*60)
    print("调度方案详细信息")
    print("="*60)
    
    for machine in sorted(machine_orders.keys()):
        print(f"\n机器 M{machine} 的加工顺序:")
        operations = machine_orders[machine]
        
        for i, op in enumerate(operations):
            start_time = start_times[op.id]
            finish_time = start_time + op.processing_time
            print(f"  {i+1}. {op.id} (工件J{op.job_id}) | "
                  f"开始时间: {start_time}, 结束时间: {finish_time}, "
                  f"加工时间: {op.processing_time}")
    
    print(f"\n总完工时间 (Makespan): {makespan}")
    print("="*60)


def run_benchmark_test(choice: str, max_iterations: int, max_attempts: int):
    """
    运行基准测试
    """
    data_files = {
        '1': ('data/ta01.txt', 'TA01'),
        '2': ('data/ta40.txt', 'TA40'), 
        '3': ('data/ta60.txt', 'TA60')
    }
    
    filename, dataset_name = data_files[choice]
    
    try:
        job_data, num_jobs, num_machines = read_benchmark_data(filename)
        compare_algorithms_on_benchmark(job_data, dataset_name, num_jobs, num_machines, max_iterations, max_attempts)
    except FileNotFoundError:
        print(f"错误: 找不到数据文件 {filename}")
        print("请确保数据文件存在于正确的路径中")
    except Exception as e:
        print(f"读取数据文件时发生错误: {e}")
        import traceback
        traceback.print_exc()

def main(choice=1, max_iterations=5000, max_attempts=3000):
    choice = str(choice).strip()

    if choice in ['1', '2', '3']:
        run_benchmark_test(choice, max_iterations, max_attempts)
    else:
        print("无效的选择，请重新运行程序")
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as patches
from collections import defaultdict, deque
import openpyxl
import random
import copy
from typing import List, Dict, Tuple, Set, Optional
import time
import pandas as pd
from Ex1 import Job, DisjunctiveGraph, ScheduleSolver, Operation

MACHINE_LIST = ['C1', 'C2', 'C3', 'C4', 'L1', 'PM1', 'NYM1', 'WYM1', 'RCL1', 'RCL2', 'WX1', 'LX1', 'LX2', 'LX3', 'Z1', 'Z2', 'DHH1', 'XQG1', 'JC1', 'QG1']


class FlexibleOperation:
    """柔性工序类"""
    def __init__(self, job_id: int, operation_id: int, machine_options: List[Tuple[str, int]], product_type: str = None):
        self.job_id = job_id
        self.operation_id = operation_id
        self.machine_options = machine_options  # [(machine_name, processing_time), ...]
        self.selected_machine = None
        self.processing_time = 0
        self.product_type = product_type or f"P{job_id}"
        self.id = f"J{job_id}O{operation_id}"
        
    def select_machine(self, machine_name: str):
        """选择加工机器"""
        for machine, time in self.machine_options:
            if machine == machine_name:
                self.selected_machine = machine
                self.processing_time = time
                return True
        return False
    
    def get_available_machines(self) -> List[str]:
        """获取可用机器列表"""
        return [machine for machine, _ in self.machine_options]
    
    def get_processing_time(self, machine_name: str) -> int:
        """获取在指定机器上的加工时间"""
        for machine, time in self.machine_options:
            if machine == machine_name:
                return time
        return float('inf')
    
    def __str__(self):
        return f"FO(J{self.job_id}O{self.operation_id}, {len(self.machine_options)} machines, product={self.product_type})"


class FlexibleJob:
    """柔性工件类"""
    def __init__(self, job_id: int, operations: List[FlexibleOperation], product_type: str = None):
        self.job_id = job_id
        self.operations = operations
        self.product_type = product_type or f"P{job_id}"
        
        # 更新所有工序的产品类型
        for op in self.operations:
            op.product_type = self.product_type
    
    def __str__(self):
        return f"FlexibleJob(J{self.job_id}, {len(self.operations)} operations, product={self.product_type})"


class FlexibleScheduleGraph:
    """柔性调度图"""
    def __init__(self, flexible_jobs: List[FlexibleJob]):
        self.flexible_jobs = flexible_jobs
        self.operations = []
        self.machine_operations = defaultdict(list)
        self.job_operations = defaultdict(list)
        self.all_machines = set()
        
        # 构建操作索引
        for job in flexible_jobs:
            for op in job.operations:
                self.operations.append(op)
                self.job_operations[job.job_id].append(op)
                for machine, _ in op.machine_options:
                    self.all_machines.add(machine)
    
    def get_machine_candidates(self, operation: FlexibleOperation) -> List[str]:
        """获取工序的候选机器"""
        return operation.get_available_machines()


class FlexibleScheduleSolver:
    """柔性调度求解器"""
    def __init__(self, graph: FlexibleScheduleGraph):
        self.graph = graph
        self.best_makespan = float('inf')
        self.best_solution = None
    
    def calculate_makespan_with_bonus(self, machine_assignment: Dict[str, str], 
                                    machine_orders: Dict[str, List[FlexibleOperation]]) -> Tuple[int, Dict[str, int]]:
        """
        计算考虑同产品连续加工优化的makespan - 修复并行执行逻辑
        """
        start_times = {}
        job_completion_times = defaultdict(int)  # 记录每个工件当前已完成的工序时间
        machine_completion_times = defaultdict(int)  # 记录每台机器的当前时间
        optimization_details = []
        
        # 创建所有工序的调度队列，按工件和工序ID排序
        all_operations = []
        for machine, operations in machine_orders.items():
            for op in operations:
                all_operations.append((machine, op))
        
        # 按工件ID和工序ID排序，确保同一工件的工序按顺序处理
        all_operations.sort(key=lambda x: (x[1].job_id, x[1].operation_id))
        
        # 逐个调度工序
        for machine, op in all_operations:
            job_id = op.job_id
            
            # 工件约束：必须等待同一工件的前序工序完成
            earliest_start_job = job_completion_times[job_id]
            
            # 机器约束：必须等待机器空闲
            earliest_start_machine = machine_completion_times[machine]
            
            # 取两者的最大值作为最早开始时间
            earliest_start = max(earliest_start_job, earliest_start_machine)
            
            # 获取基础加工时间
            base_time = op.get_processing_time(machine_assignment[op.id])
            processing_time = base_time
            
            # 检查同产品连续加工优化
            if machine in machine_orders and len(machine_orders[machine]) > 0:
                # 找到在同一机器上的前一个工序
                machine_ops = machine_orders[machine]
                current_index = -1
                for i, machine_op in enumerate(machine_ops):
                    if machine_op.id == op.id:
                        current_index = i
                        break
                
                if current_index > 0:
                    prev_op = machine_ops[current_index - 1]
                    # 检查是否满足连续加工条件
                    prev_finish_time = start_times.get(prev_op.id, 0) + prev_op.get_processing_time(machine_assignment[prev_op.id])
                    
                    if (prev_op.product_type == op.product_type and 
                        earliest_start == machine_completion_times[machine] and
                        prev_finish_time == machine_completion_times[machine]):
                        # 应用25%的时间优化
                        processing_time = max(1, int(base_time * 0.75))
                        optimization_details.append({
                            'machine': machine,
                            'operation': op.id,
                            'product': op.product_type,
                            'original_time': base_time,
                            'optimized_time': processing_time,
                            'savings': base_time - processing_time
                        })
            
            # 设置开始时间和完成时间
            start_times[op.id] = earliest_start
            completion_time = earliest_start + processing_time
            
            # 更新工件完成时间（该工件的最新工序完成时间）
            job_completion_times[job_id] = completion_time
            
            # 更新机器完成时间
            machine_completion_times[machine] = completion_time
        
        makespan = max(machine_completion_times.values()) if machine_completion_times else 0
        self._last_optimization_details = optimization_details
        
        return makespan, start_times
    
    def _get_job_constraint_time(self, operation: FlexibleOperation, 
                                start_times: Dict[str, int], 
                                machine_assignment: Dict[str, str]) -> int:
        """
        获取工件内工序约束的最早开始时间
        """
        job_id = operation.job_id
        operation_id = operation.operation_id
        
        # 找到同一工件的所有工序
        job_operations = []
        for op in self.graph.operations:
            if op.job_id == job_id:
                job_operations.append(op)
        
        # 按工序ID排序
        job_operations.sort(key=lambda x: x.operation_id)
        
        # 找到当前工序的位置
        current_index = -1
        for i, op in enumerate(job_operations):
            if op.operation_id == operation_id:
                current_index = i
                break
        
        # 如果是第一个工序，无前序约束
        if current_index <= 0:
            return 0
        
        # 计算前序工序的完成时间
        max_predecessor_finish = 0
        for i in range(current_index):
            pred_op = job_operations[i]
            if pred_op.id in start_times:
                pred_start = start_times[pred_op.id]
                pred_duration = pred_op.get_processing_time(machine_assignment[pred_op.id])
                pred_finish = pred_start + pred_duration
                max_predecessor_finish = max(max_predecessor_finish, pred_finish)
        
        return max_predecessor_finish
    
    def _initialize_population(self, size: int) -> List[Tuple[Dict[str, str], Dict[str, List]]]:
        """初始化种群"""
        population = []
        
        for _ in range(size):
            # 随机机器分配
            machine_assignment = {}
            for op in self.graph.operations:
                available_machines = op.get_available_machines()
                machine_assignment[op.id] = random.choice(available_machines)
            
            # 基于机器分配生成工序顺序
            machine_orders = self._generate_random_order(machine_assignment)
            
            population.append((machine_assignment, machine_orders))
        
        return population
    
    def _generate_random_order(self, machine_assignment: Dict[str, str]) -> Dict[str, List]:
        """基于机器分配生成随机工序顺序，保持同一工件内工序的顺序约束"""
        machine_operations = defaultdict(list)
        
        # 将工序分配到对应机器
        for op in self.graph.operations:
            machine = machine_assignment[op.id]
            machine_operations[machine].append(op)
        
        # 对每台机器上的工序进行排序，确保同一工件的工序按序号排列
        for machine in machine_operations:
            # 先按工件ID分组，再按工序ID排序
            machine_operations[machine].sort(key=lambda x: (x.job_id, x.operation_id))
            
            # 在保持工件内顺序的前提下，随机排列不同工件的工序
            jobs_ops = defaultdict(list)
            for op in machine_operations[machine]:
                jobs_ops[op.job_id].append(op)
            
            # 随机交织不同工件的工序序列
            result = []
            job_indices = {job_id: 0 for job_id in jobs_ops.keys()}
            remaining_jobs = list(jobs_ops.keys())
            
            while remaining_jobs:
                # 随机选择一个工件
                job_id = random.choice(remaining_jobs)
                job_ops = jobs_ops[job_id]
                idx = job_indices[job_id]
                
                if idx < len(job_ops):
                    result.append(job_ops[idx])
                    job_indices[job_id] += 1
                
                # 如果该工件的工序全部安排完，从候选列表中移除
                if job_indices[job_id] >= len(job_ops):
                    remaining_jobs.remove(job_id)
            
            machine_operations[machine] = result
        
        return dict(machine_operations)
    
    def _roulette_selection(self, population: List, fitness_scores: List[float]):
        """轮盘赌选择"""
        total_fitness = sum(fitness_scores)
        if total_fitness == 0:
            return random.choice(population)
        
        pick = random.uniform(0, total_fitness)
        current = 0
        for i, fitness in enumerate(fitness_scores):
            current += fitness
            if current >= pick:
                return population[i]
        return population[-1]
    
    def _crossover(self, parent1: Tuple, parent2: Tuple) -> Tuple[Tuple, Tuple]:
        """交叉操作 - 保持工件内工序顺序"""
        assignment1, orders1 = parent1
        assignment2, orders2 = parent2
        
        # 机器分配交叉
        child_assignment1 = assignment1.copy()
        child_assignment2 = assignment2.copy()
        
        # 随机选择一部分工序进行交换
        operations = list(assignment1.keys())
        crossover_point = len(operations) // 2
        random.shuffle(operations)
        
        for op_id in operations[:crossover_point]:
            child_assignment1[op_id] = assignment2[op_id]
            child_assignment2[op_id] = assignment1[op_id]
        
        # 基于新的机器分配重新生成工序顺序（保持约束）
        child_orders1 = self._generate_random_order(child_assignment1)
        child_orders2 = self._generate_random_order(child_assignment2)
        
        return (child_assignment1, child_orders1), (child_assignment2, child_orders2)
    
    def _mutate(self, individual: Tuple) -> Tuple:
        """变异操作 - 保持工件内工序顺序"""
        assignment, orders = individual
        new_assignment = assignment.copy()
        
        # 随机改变几个工序的机器分配
        operations = list(assignment.keys())
        num_mutations = max(1, len(operations) // 20)  # 减少变异强度
        
        for _ in range(num_mutations):
            op_id = random.choice(operations)
            # 找到对应的工序对象
            operation = None
            for op in self.graph.operations:
                if op.id == op_id:
                    operation = op
                    break
            
            if operation:
                available_machines = operation.get_available_machines()
                if len(available_machines) > 1:  # 只有多个选择时才变异
                    current_machine = new_assignment[op_id]
                    available_machines = [m for m in available_machines if m != current_machine]
                    if available_machines:
                        new_assignment[op_id] = random.choice(available_machines)
        
        # 重新生成工序顺序（保持约束）
        new_orders = self._generate_random_order(new_assignment)
        
        return (new_assignment, new_orders)

    def genetic_algorithm(self, population_size: int = 100, generations: int = 500,
                         mutation_rate: float = 0.1, crossover_rate: float = 0.8) -> Tuple[int, Dict, Dict]:
        """
        遗传算法求解FJSP
        """
        print(f"开始遗传算法求解 (种群大小: {population_size}, 代数: {generations})")
        
        # 初始化种群
        population = self._initialize_population(population_size)
        
        best_makespan = float('inf')
        best_solution = None
        best_orders = None
        generation_bests = []
        
        for generation in range(generations):
            # 评估种群适应度
            fitness_scores = []
            for individual in population:
                machine_assignment, machine_orders = individual
                makespan, _ = self.calculate_makespan_with_bonus(machine_assignment, machine_orders)
                fitness_scores.append(1.0 / (makespan + 1))  # 适应度为倒数
                
                if makespan < best_makespan:
                    best_makespan = makespan
                    best_solution = machine_assignment.copy()
                    best_orders = {m: ops.copy() for m, ops in machine_orders.items()}
            
            generation_bests.append(best_makespan)
            
            # 选择、交叉、变异
            new_population = []
            
            # 保留最优个体（精英策略）
            best_idx = fitness_scores.index(max(fitness_scores))
            new_population.append(population[best_idx])
            
            # 生成新个体
            while len(new_population) < population_size:
                # 轮盘赌选择
                parent1 = self._roulette_selection(population, fitness_scores)
                parent2 = self._roulette_selection(population, fitness_scores)
                
                # 交叉
                if random.random() < crossover_rate:
                    child1, child2 = self._crossover(parent1, parent2)
                else:
                    child1, child2 = parent1, parent2
                
                # 变异
                if random.random() < mutation_rate:
                    child1 = self._mutate(child1)
                if random.random() < mutation_rate:
                    child2 = self._mutate(child2)
                
                new_population.extend([child1, child2])
            
            population = new_population[:population_size]
            
            # if generation % 50 == 0:
            print(f"第 {generation} 代，最佳makespan: {best_makespan}")
        
        print(f"遗传算法完成，最佳makespan: {best_makespan}")
        
        # 绘制收敛曲线
        self._plot_convergence(generation_bests)
        
        return best_makespan, best_solution, best_orders
    
    def _plot_convergence(self, generation_bests: List[int]):
        """绘制算法收敛曲线"""
        plt.figure(figsize=(10, 6))
        plt.plot(generation_bests, 'b-', linewidth=2, alpha=0.8)
        plt.xlabel('Generation', fontsize=12)
        plt.ylabel('Best Makespan', fontsize=12)
        plt.title('遗传算法收敛曲线', fontsize=14)
        plt.grid(True, alpha=0.3)
        plt.tight_layout()
        plt.show()
    
    def analyze_optimization_potential(self, machine_assignment: Dict[str, str], 
                                     machine_orders: Dict[str, List[FlexibleOperation]]) -> Dict:
        """
        分析同产品连续加工优化的潜力
        """
        analysis = {
            'total_operations': 0,
            'potential_optimizations': 0,
            'actual_optimizations': 0,
            'time_saved': 0,
            'product_analysis': defaultdict(lambda: {'operations': 0, 'optimizations': 0, 'savings': 0})
        }
        
        # 重新计算以获取优化详情
        self._last_optimization_details = []
        makespan, start_times = self.calculate_makespan_with_bonus(machine_assignment, machine_orders)
        
        # 分析每台机器上的工序序列
        for machine, operations in machine_orders.items():
            prev_product = None
            for i, op in enumerate(operations):
                analysis['total_operations'] += 1
                analysis['product_analysis'][op.product_type]['operations'] += 1
                
                if prev_product == op.product_type:
                    analysis['potential_optimizations'] += 1
                
                prev_product = op.product_type
        
        # 统计实际优化
        for opt in self._last_optimization_details:
            analysis['actual_optimizations'] += 1
            analysis['time_saved'] += opt['savings']
            analysis['product_analysis'][opt['product']]['optimizations'] += 1
            analysis['product_analysis'][opt['product']]['savings'] += opt['savings']
        
        return analysis


def validate_job_data(job_data: List[Tuple[List[int], List[int]]], num_jobs: int, num_machines: int) -> List[Job]:
    """
    验证工件数据并创建工件对象
    
    Args:
        job_data: 工件数据列表
        num_jobs: 工件数量
        num_machines: 机器数量
        
    Returns:
        List[Job]: 验证后的工件对象列表
    """
    print("正在验证数据完整性...")
    
    # 检查数据合理性
    for i, (machines, times) in enumerate(job_data):
        if len(machines) != len(times):
            raise ValueError(f"工件 {i+1} 的机器和时间数据长度不匹配")
        
        if any(t <= 0 for t in times):
            print(f"警告: 工件 {i+1} 包含非正数加工时间")
        
        if any(m < 1 for m in machines):
            print(f"警告: 工件 {i+1} 包含无效机器编号")
    
    # 创建工件对象
    jobs = []
    for i, (machines, times) in enumerate(job_data):
        try:
            job = Job(i + 1, machines, times)
            jobs.append(job)
            if i < 3:
                print(f"工件 {i+1}: 机器序列长度={len(machines)}, 加工时间范围=[{min(times)}, {max(times)}]")
        except Exception as e:
            raise ValueError(f"创建工件 {i+1} 时发生错误: {e}")
    
    return jobs

def build_disjunctive_graph_and_solver(jobs: List[Job]) -> Tuple[DisjunctiveGraph, ScheduleSolver]:
    """
    构建析取图和求解器
    
    Args:
        jobs: 工件列表
        
    Returns:
        Tuple[DisjunctiveGraph, ScheduleSolver]: 析取图和求解器
    """
    try:
        graph = DisjunctiveGraph(jobs)
        solver = ScheduleSolver(graph)
        
        print(f"\n析取图构建完成:")
        print(f"- 工序总数: {len(graph.operations)}")
        print(f"- 连接弧数量: {len(graph.conjunctive_arcs)}")
        print(f"- 析取弧数量: {len(graph.disjunctive_arcs)}")
        
        # 验证机器操作分布
        print(f"- 各机器上的工序数量:")
        for machine in sorted(graph.machine_operations.keys()):
            ops_count = len(graph.machine_operations[machine])
            print(f"  机器 M{machine}: {ops_count} 个工序")
            
        return graph, solver
        
    except Exception as e:
        raise ValueError(f"构建析取图时发生错误: {e}")

def solve_and_visualize(solver: ScheduleSolver, dataset_name: str, num_jobs: int, max_iterations: int, max_attempts: int) -> Dict[str, Tuple]:
    """
    求解调度问题并可视化结果
    
    Args:
        solver: 调度求解器
        dataset_name: 数据集名称
        num_jobs: 工件数量
        max_iterations: 最大迭代次数
        max_attempts: 最大尝试次数
        
    Returns:
        Dict[str, Tuple]: 求解结果
    """
    print(f"\n运行随机搜索算法 (最大迭代次数: {max_iterations})...")
    
    try:
        start_time = time.time()
        best_makespan_rs, best_orders_rs = solver.random_search(max_iterations=max_iterations, max_attempts=max_attempts)
        rs_time = time.time() - start_time
        
        if best_makespan_rs != float('inf'):
            print(f"随机搜索完成: Makespan = {best_makespan_rs}, 时间 = {rs_time:.3f}s")
            
            # 生成甘特图
            print(f"\n正在为 {dataset_name} 生成甘特图...")
            try:
                _, start_times = solver.calculate_makespan(best_orders_rs)
                visualize_schedule(best_orders_rs, start_times, 
                                    f"{dataset_name} - 随机搜索 (Makespan = {best_makespan_rs})")
                
                # 显示调度详情
                if num_jobs <= 10:  # 只对小规模问题显示详细信息
                    print_schedule_details(best_orders_rs, start_times, best_makespan_rs)
                    
            except Exception as e:
                print(f"生成甘特图时出错: {e}")
                
            return {'随机搜索': (best_makespan_rs, rs_time)}
        else:
            print(f"随机搜索未找到有效解")
            return {'随机搜索': (None, None)}
            
    except Exception as e:
        print(f"随机搜索执行出错: {e}")
        return {'随机搜索': (None, None)}

def read_real_case_data(filename: str) -> Tuple[List[FlexibleJob], Dict[str, str]]:
    """
    读取实际案例Excel数据
    
    Args:
        filename: Excel文件路径
        
    Returns:
        (jobs, machine_mapping): 柔性工件列表和机器映射表
    """

    # 使用openpyxl读取Excel文件
    workbook = openpyxl.load_workbook(filename)
    worksheet = workbook.active
    
    print(f"读取Excel文件: {filename}")
    print(f"工作表名称: {worksheet.title}")
    print(f"数据范围: {worksheet.max_row} 行 x {worksheet.max_column} 列")
    
    # 解析数据结构
    jobs = []
    machine_mapping = {}
    
    # 从第5行开始读取数据
    current_job_id = 1
    current_job_operations = []
    current_quantity = 1
    
    # 机器列名映射 (F列开始对应所有机器，E列是工序序号)
    machine_columns = MACHINE_LIST
    part_name = ""
    for row in range(6, worksheet.max_row + 1):  # 从第6行开始

        # 读取各列数据
        cur_part_name = worksheet.cell(row=row, column=2).value  # B列：零件名称
        
        cur_quantity = worksheet.cell(row=row, column=3).value   # C列：数量
        feature = worksheet.cell(row=row, column=4).value    # D列：特征
        operation_seq = worksheet.cell(row=row, column=5).value  # E列：工序序号
        feature = str(feature).strip() if feature else ""
        
        # 处理数量
        quantity = int(float(str(cur_quantity).strip())) if cur_quantity else quantity
        
        # 处理工序序号
        operation_seq = int(float(str(operation_seq).strip())) if operation_seq else 0

        # 检查是否开始新工件（B列出现新值）
        if cur_part_name is not None and part_name != cur_part_name:
            
            # 创建前一个工件（根据数量创建多个相同工件）
            if current_job_operations:
                for i in range(current_quantity):
                    print(f"添加工件 {current_job_id}，零件={part_name}, 工序数={len(current_job_operations)}")
                    job = FlexibleJob(current_job_id, copy.deepcopy(current_job_operations), part_name)
                    jobs.append(job)
                    current_job_id += 1
            part_name = str(cur_part_name).strip() if cur_part_name else part_name
            current_job_operations = []

        current_quantity = quantity
        
        # 解析机器选项和时间（从F列开始）
        machine_options = []
        
        # 读取F列到后续列的机器时间数据
        for i, machine_name in enumerate(machine_columns):
            col_idx = 6 + i  # F列开始（索引6）
            cell_value = worksheet.cell(row=row, column=col_idx).value
            
            if cell_value is not None and str(cell_value).strip() != '':
                try:
                    time_val = int(float(str(cell_value).strip()))
                    if time_val > 0:  # 只添加正数时间
                        machine_options.append((machine_name, time_val))
                except (ValueError, TypeError):
                    continue
        
        # 如果有有效的机器选项，创建工序
        if machine_options and operation_seq > 0:
            # 使用Excel中的工序序号
            operation = FlexibleOperation(current_job_id, operation_seq, machine_options, part_name)
            current_job_operations.append(operation)
    
    # 添加最后一个工件（根据数量创建多个相同工件）
    if current_job_operations:
        for i in range(current_quantity):
            job = FlexibleJob(current_job_id, copy.deepcopy(current_job_operations), part_name)
            jobs.append(job)
            current_job_id += 1

    # 创建机器映射
    all_machines = set()
    for job in jobs:
        for op in job.operations:
            for machine, _ in op.machine_options:
                all_machines.add(machine)
    
    machine_mapping = {machine: machine for machine in sorted(all_machines)}
    
    print(f"\n解析完成:")
    print(f"工件数量: {len(jobs)}")
    print(f"机器数量: {len(machine_mapping)}")
    print(f"机器列表: {sorted(machine_mapping.keys())}")
    
    workbook.close()
    return jobs, machine_mapping


def visualize_flexible_schedule(machine_assignment: Dict[str, str],
                               machine_orders: Dict[str, List[FlexibleOperation]], 
                               start_times: Dict[str, int], 
                               title: str = "柔性作业车间调度甘特图"):
    """
    可视化柔性调度方案的甘特图 - 修复显示问题
    """
    if not machine_orders or not start_times:
        print("无效的调度数据，无法生成甘特图")
        return
        
    fig, ax = plt.subplots(figsize=(20, 12))
    
    # 扩展颜色映射
    colors = [
        '#FF6B6B', '#4ECDC4', '#45B7D1', '#96CEB4', '#FFEAA7', 
        '#DDA0DD', '#98D8C8', '#F7DC6F', '#BB8FCE', '#85C1E9',
        '#F8C471', '#82E0AA', '#F1948A', '#85C1E9', '#D2B4DE',
        '#AED6F1', '#A3E4D7', '#D5DBDB', '#FADBD8', '#D1F2EB',
        '#FCF3CF', '#EBDEF0', '#EAF2F8', '#E8F8F5', '#FDF2E9'
    ]
    
    # 为不同产品分配颜色
    product_colors = {}
    job_colors = {}
    
    machines = sorted(machine_orders.keys())
    y_positions = {machine: i for i, machine in enumerate(machines)}
    
    # 计算最大时间
    max_time = 0
    for ops in machine_orders.values():
        for op in ops:
            if op.id in start_times:
                finish_time = start_times[op.id] + op.get_processing_time(machine_assignment[op.id])
                max_time = max(max_time, finish_time)
    
    print(f"甘特图信息:")
    print(f"  机器数量: {len(machines)}")
    print(f"  总工序数: {sum(len(ops) for ops in machine_orders.values())}")
    print(f"  最大时间: {max_time}")
    
    # 绘制甘特图
    for machine, operations in machine_orders.items():
        y_pos = y_positions[machine]
        print(f"  机器 {machine}: {len(operations)} 个工序")
        
        for i, op in enumerate(operations):
            # 为产品类型分配颜色
            if op.product_type not in product_colors:
                product_colors[op.product_type] = colors[len(product_colors) % len(colors)]
            
            # 为工件分配相同产品的颜色
            if op.job_id not in job_colors:
                base_color = product_colors[op.product_type]
                job_colors[op.job_id] = base_color
            
            if op.id in start_times:
                start_time = start_times[op.id]
                duration = op.get_processing_time(machine_assignment[op.id])
                
                print(f"    工序 {op.id}: 开始={start_time}, 持续={duration}, 结束={start_time + duration}")
                
                # 绘制工序矩形
                rect = patches.Rectangle((start_time, y_pos - 0.35), duration, 0.7,
                                       linewidth=1.5, edgecolor='white', 
                                       facecolor=job_colors[op.job_id], alpha=0.8)
                ax.add_patch(rect)
                
                # 添加工序标签
                if duration > max_time * 0.01:  # 降低显示文字的阈值
                    ax.text(start_time + duration/2, y_pos, f'J{op.job_id}O{op.operation_id}',
                           ha='center', va='center', fontsize=9, fontweight='bold', color='black')
    
    # 设置图表属性
    ax.set_xlabel('Time (时间单位)', fontsize=16, fontweight='bold')
    ax.set_ylabel('Machine (机器)', fontsize=16, fontweight='bold')
    ax.set_title(title, fontsize=18, fontweight='bold', pad=20)
    
    # 设置y轴
    ax.set_yticks(range(len(machines)))
    ax.set_yticklabels([f'{m}' for m in machines], fontsize=12)
    ax.set_ylim(-0.5, len(machines) - 0.5)
    
    # 设置x轴
    if max_time > 0:
        ax.set_xlim(0, max_time + max_time * 0.05)
        # 添加时间刻度
        time_ticks = range(0, int(max_time) + 1, max(1, int(max_time) // 20))
        ax.set_xticks(time_ticks)
    ax.tick_params(axis='x', labelsize=12)
    ax.tick_params(axis='y', labelsize=12)
    
    # 添加网格
    ax.grid(True, alpha=0.3, linestyle='-', linewidth=0.5)
    ax.set_axisbelow(True)
    
    # 添加产品图例
    legend_elements = []
    for product_type in sorted(product_colors.keys()):
        legend_elements.append(
            patches.Patch(facecolor=product_colors[product_type], 
                         label=f'{product_type}', alpha=0.8,
                         edgecolor='white', linewidth=1)
        )
    
    if legend_elements:
        ax.legend(handles=legend_elements, loc='center left', 
                 bbox_to_anchor=(1.02, 0.5), fontsize=12,
                 title='产品类型', title_fontsize=14)
    
    # 添加makespan标记线
    if max_time > 0:
        ax.axvline(x=max_time, color='red', linestyle='--', linewidth=3, alpha=0.8)
        ax.text(max_time + max_time * 0.01, len(machines) * 0.95, 
               f'Makespan = {max_time}', 
               fontsize=14, fontweight='bold', color='red',
               bbox=dict(boxstyle="round,pad=0.3", facecolor="white", alpha=0.8))
    
    # 添加机器利用率信息
    machine_utilization = {}
    for machine in machines:
        total_work_time = 0
        if machine in machine_orders:
            for op in machine_orders[machine]:
                if op.id in start_times:
                    total_work_time += op.get_processing_time(machine_assignment[op.id])
        utilization = (total_work_time / max_time * 100) if max_time > 0 else 0
        machine_utilization[machine] = utilization
    
    # 在图表上显示利用率
    for i, machine in enumerate(machines):
        util = machine_utilization[machine]
        ax.text(-max_time * 0.02, i, f'{util:.1f}%', 
               ha='right', va='center', fontsize=10, 
               bbox=dict(boxstyle="round,pad=0.2", facecolor="lightblue", alpha=0.7))
    
    ax.set_facecolor('#FAFAFA')
    plt.tight_layout()
    plt.show()
    
    # 打印详细的调度统计
    print(f"\n机器利用率统计:")
    for machine in sorted(machines):
        util = machine_utilization[machine]
        work_time = sum(op.get_processing_time(machine_assignment[op.id]) 
                       for op in machine_orders.get(machine, []) 
                       if op.id in start_times)
        print(f"  {machine}: 工作时间={work_time}, 利用率={util:.1f}%")

def solve_real_case(max_iterations: int = 5000):
    """求解实际案例"""
    print("开始处理实际案例数据...")
    
    # 读取Excel数据
    filename = "data/real_case.xlsx"
    flexible_jobs, machine_mapping = read_real_case_data(filename)
    
    if not flexible_jobs:
        print("未能成功读取数据，请检查Excel文件格式")
        return
    
    print(f"\n构建柔性调度图...")
    graph = FlexibleScheduleGraph(flexible_jobs)
    solver = FlexibleScheduleSolver(graph)
    
    print(f"工序总数: {len(graph.operations)}")
    print(f"机器总数: {len(graph.all_machines)}")
    print(f"机器列表: {sorted(graph.all_machines)}")
    
    # 显示问题统计信息
    total_machine_options = sum(len(op.machine_options) for op in graph.operations)
    avg_flexibility = total_machine_options / len(graph.operations) if graph.operations else 0
    print(f"平均柔性度: {avg_flexibility:.2f}")
    
    # 显示产品类型统计
    product_stats = defaultdict(int)
    for job in flexible_jobs:
        product_stats[job.product_type] += 1
    
    print(f"\n产品类型统计:")
    for product, count in sorted(product_stats.items()):
        print(f"  {product}: {count} 个工件")
    
    # 运行遗传算法
    print(f"\n开始求解柔性作业车间调度问题...")
    start_time = time.time()
    
    best_makespan, best_assignment, best_orders = solver.genetic_algorithm(
        population_size=100, 
        generations=max_iterations,
        mutation_rate=0.12,
        crossover_rate=0.88
    )
    
    solve_time = time.time() - start_time
    
    if best_makespan != float('inf'):
        print(f"\n求解完成!")
        print(f"最优makespan: {best_makespan}")
        print(f"求解时间: {solve_time:.3f}秒")
        
        # 分析优化效果
        optimization_analysis = solver.analyze_optimization_potential(best_assignment, best_orders)
        
        print(f"\n同产品连续加工优化分析:")
        print(f"  总工序数: {optimization_analysis['total_operations']}")
        print(f"  潜在优化机会: {optimization_analysis['potential_optimizations']}")
        print(f"  实际优化次数: {optimization_analysis['actual_optimizations']}")
        print(f"  节省总时间: {optimization_analysis['time_saved']}")
        print(f"  优化率: {optimization_analysis['actual_optimizations']/optimization_analysis['potential_optimizations']*100:.1f}%" if optimization_analysis['potential_optimizations'] > 0 else "  优化率: 0%")
        
        print(f"\n各产品优化详情:")
        for product, stats in optimization_analysis['product_analysis'].items():
            if stats['operations'] > 0:
                print(f"  {product}: {stats['operations']}工序, {stats['optimizations']}次优化, 节省{stats['savings']}时间")
        
        # 计算详细调度信息
        _, start_times = solver.calculate_makespan_with_bonus(best_assignment, best_orders)
        
        # 生成甘特图
        visualize_flexible_schedule(
            best_assignment, best_orders, start_times,
            f"柔性作业车间调度结果 (Makespan = {best_makespan})"
        )
        
        # 显示机器分配统计
        print(f"\n机器分配统计:")
        machine_usage = defaultdict(int)
        machine_time_usage = defaultdict(int)
        for op_id, machine in best_assignment.items():
            machine_usage[machine] += 1
            # 找到对应的操作并计算时间
            for op in graph.operations:
                if op.id == op_id:
                    machine_time_usage[machine] += op.get_processing_time(machine)
                    break
        
        for machine in sorted(machine_usage.keys()):
            utilization = machine_time_usage[machine] / best_makespan * 100 if best_makespan > 0 else 0
            print(f"  {machine}: {machine_usage[machine]}工序, {machine_time_usage[machine]}时间, 利用率{utilization:.1f}%")
        
        # 显示机器负载平衡度
        if machine_time_usage:
            avg_load = sum(machine_time_usage.values()) / len(machine_time_usage)
            load_variance = sum((load - avg_load)**2 for load in machine_time_usage.values()) / len(machine_time_usage)
            print(f"\n负载平衡分析:")
            print(f"  平均机器负载: {avg_load:.1f}")
            print(f"  负载方差: {load_variance:.1f}")
            print(f"  负载均衡系数: {1/(1+load_variance/avg_load**2):.3f}" if avg_load > 0 else "  负载均衡系数: N/A")
        
    else:
        print("未找到可行解")


def visualize_schedule(machine_orders: Dict[int, List[Operation]], 
                      start_times: Dict[str, int], 
                      title: str = "作业车间调度甘特图"):
    """
    可视化调度方案的甘特图
    
    Args:
        machine_orders: 每台机器的工序顺序
        start_times: 各工序的开始时间
        title: 图表标题
    """
    # 检查输入数据
    if not machine_orders or not start_times:
        print("无效的调度数据，无法生成甘特图")
        return
        
    fig, ax = plt.subplots(figsize=(14, 8))
    
    # 扩展颜色映射（为不同工件分配不同颜色）
    colors = [
        '#FF6B6B', '#4ECDC4', '#45B7D1', '#96CEB4', '#FFEAA7', 
        '#DDA0DD', '#98D8C8', '#F7DC6F', '#BB8FCE', '#85C1E9',
        '#F8C471', '#82E0AA', '#F1948A', '#85C1E9', '#D2B4DE',
        '#AED6F1', '#A3E4D7', '#D5DBDB', '#FADBD8', '#D1F2EB',
        '#FCF3CF', '#EBDEF0', '#EAF2F8', '#E8F8F5', '#FDF2E9'
    ]
    job_colors = {}
    
    machines = sorted(machine_orders.keys())
    y_positions = {machine: i for i, machine in enumerate(machines)}
    
    # 计算最大时间用于设置x轴
    max_time = max(start_times[op.id] + op.processing_time 
                  for ops in machine_orders.values() for op in ops)
    
    # 绘制甘特图
    for machine, operations in machine_orders.items():
        y_pos = y_positions[machine]
        
        for op in operations:
            job_id = op.job_id
            if job_id not in job_colors:
                job_colors[job_id] = colors[len(job_colors) % len(colors)]
            
            start_time = start_times[op.id]
            duration = op.processing_time
            
            # 绘制工序矩形 - 去掉文字标签
            rect = patches.Rectangle((start_time, y_pos - 0.35), duration, 0.7,
                                   linewidth=1.5, edgecolor='white', 
                                   facecolor=job_colors[job_id], alpha=0.8)
            ax.add_patch(rect)
    
    # 设置图表属性
    ax.set_xlabel('Time', fontsize=16, fontweight='bold')
    ax.set_ylabel('Machine', fontsize=16, fontweight='bold')
    ax.set_title(title, fontsize=18, fontweight='bold', pad=20)
    
    # 设置y轴
    ax.set_yticks(range(len(machines)))
    ax.set_yticklabels([f'M{m}' for m in machines], fontsize=14)
    ax.set_ylim(-0.5, len(machines) - 0.5)
    
    # 设置x轴
    ax.set_xlim(0, max_time + max_time * 0.05)  # 添加5%的边距
    ax.tick_params(axis='x', labelsize=12)
    ax.tick_params(axis='y', labelsize=12)
    
    # 添加网格
    ax.grid(True, alpha=0.3, linestyle='-', linewidth=0.5)
    ax.set_axisbelow(True)
    
    # 添加图例 - 显示工件信息
    legend_elements = []
    for job_id in sorted(job_colors.keys()):
        legend_elements.append(
            patches.Patch(facecolor=job_colors[job_id], 
                         label=f'Job {job_id}', alpha=0.8,
                         edgecolor='white', linewidth=1)
        )
    
    # 将图例放在图表右侧
    ax.legend(handles=legend_elements, loc='center left', 
             bbox_to_anchor=(1.02, 0.5), fontsize=12,
             title='Jobs', title_fontsize=14)
    
    # 添加makespan标记线
    ax.axvline(x=max_time, color='red', linestyle='--', linewidth=3, alpha=0.8)
    ax.text(max_time + max_time * 0.01, len(machines) * 0.95, 
           f'Makespan = {max_time}', 
           fontsize=14, fontweight='bold', color='red',
           bbox=dict(boxstyle="round,pad=0.3", facecolor="white", alpha=0.8))
    
    # 设置背景色
    ax.set_facecolor('#FAFAFA')
    
    plt.tight_layout()
    plt.show()

def print_schedule_details(machine_orders: Dict[int, List[Operation]], 
                          start_times: Dict[str, int], 
                          makespan: int):
    """
    打印调度方案的详细信息
    
    Args:
        machine_orders: 每台机器的工序顺序
        start_times: 各工序的开始时间
        makespan: 最大完工时间
    """
    print("\n" + "="*60)
    print("调度方案详细信息")
    print("="*60)
    
    for machine in sorted(machine_orders.keys()):
        print(f"\n机器 M{machine} 的加工顺序:")
        operations = machine_orders[machine]
        
        for i, op in enumerate(operations):
            start_time = start_times[op.id]
            finish_time = start_time + op.processing_time
            print(f"  {i+1}. {op.id} (工件J{op.job_id}) | "
                  f"开始时间: {start_time}, 结束时间: {finish_time}, "
                  f"加工时间: {op.processing_time}")
    
    print(f"\n总完工时间 (Makespan): {makespan}")
    print("="*60)


if __name__ == "__main__":
    # 设置中文字体支持
    plt.rcParams['font.sans-serif'] = ['SimHei', 'DejaVu Sans']
    plt.rcParams['axes.unicode_minus'] = False
    main(choice=2, max_iterations=10000, max_attempts=3000000)
    # 运行柔性作业车间调度实际案例
    solve_real_case(max_iterations=500)
   
