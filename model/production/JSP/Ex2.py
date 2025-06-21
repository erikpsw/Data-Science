import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as patches
from collections import defaultdict, deque
import openpyxl
import random
import copy
from typing import List, Dict, Tuple, Set, Optional
import time
import pandas as pd
from Ex1 import Job, DisjunctiveGraph, ScheduleSolver, Operation

MACHINE_LIST = ['C1', 'C2', 'C3', 'C4', 'L1', 'PM1', 'NYM1', 'WYM1', 'RCL1', 'RCL2', 'WX1', 'LX1', 'LX2', 'LX3', 'Z1', 'Z2', 'DHH1', 'XQG1', 'JC1', 'QG1']


class FlexibleOperation:
    """柔性工序类"""
    def __init__(self, job_id: int, operation_id: int, machine_options: List[Tuple[str, int]], product_type: str = None):
        self.job_id = job_id
        self.operation_id = operation_id
        self.machine_options = machine_options  # [(machine_name, processing_time), ...]
        self.selected_machine = None
        self.processing_time = 0
        self.product_type = product_type or f"P{job_id}"
        self.id = f"J{job_id}O{operation_id}"
        
    def select_machine(self, machine_name: str):
        """选择加工机器"""
        for machine, time in self.machine_options:
            if machine == machine_name:
                self.selected_machine = machine
                self.processing_time = time
                return True
        return False
    
    def get_available_machines(self) -> List[str]:
        """获取可用机器列表"""
        return [machine for machine, _ in self.machine_options]
    
    def get_processing_time(self, machine_name: str) -> int:
        """获取在指定机器上的加工时间"""
        for machine, time in self.machine_options:
            if machine == machine_name:
                return time
        return float('inf')
    
    def __str__(self):
        return f"FO(J{self.job_id}O{self.operation_id}, {len(self.machine_options)} machines, product={self.product_type})"


class FlexibleJob:
    """柔性工件类"""
    def __init__(self, job_id: int, operations: List[FlexibleOperation], product_type: str = None):
        self.job_id = job_id
        self.operations = operations
        self.product_type = product_type or f"P{job_id}"
        
        # 更新所有工序的产品类型
        for op in self.operations:
            op.product_type = self.product_type
    
    def __str__(self):
        return f"FlexibleJob(J{self.job_id}, {len(self.operations)} operations, product={self.product_type})"


class FlexibleSchedule:
    """柔性调度"""
    def __init__(self, flexible_jobs: List[FlexibleJob]):
        self.flexible_jobs = flexible_jobs
        self.operations = []
        self.machine_operations = defaultdict(list)
        self.job_operations = defaultdict(list)
        self.all_machines = set()
        
        # 构建操作索引
        for job in flexible_jobs:
            for op in job.operations:
                self.operations.append(op)
                self.job_operations[job.job_id].append(op)
                for machine, _ in op.machine_options:
                    self.all_machines.add(machine)
    
    def get_machine_candidates(self, operation: FlexibleOperation) -> List[str]:
        """获取工序的候选机器"""
        return operation.get_available_machines()


class FlexibleScheduleSolver:
    """柔性调度求解器"""
    def __init__(self, schedule: FlexibleSchedule):
        self.schedule = schedule
        self.best_makespan = float('inf')
        self.best_solution = None
    
    def calculate_makespan(self, machine_assignment: Dict[str, str], 
                          machine_orders: Dict[str, List[FlexibleOperation]]) -> Tuple[int, Dict[str, int]]:
        """
        计算makespan - 简化版本，只考虑基本约束
        """
        start_times = {}
        job_completion_times = defaultdict(int)  # 记录每个工件当前已完成的工序时间
        machine_completion_times = defaultdict(int)  # 记录每台机器的当前时间
        
        # 按机器分组的工序，保持原有顺序
        scheduled_operations = []
        for machine, operations in machine_orders.items():
            for position, op in enumerate(operations):
                scheduled_operations.append((machine, op, position))

        # 按工件ID和工序ID排序，确保工件内工序约束
        scheduled_operations.sort(key=lambda x: (x[1].job_id, x[1].operation_id))
        # 逐个调度工序
        for machine, op, original_position in scheduled_operations:
            job_id = op.job_id
            
            # 工件约束：必须等待同一工件的前序工序完成
            earliest_start_job = job_completion_times[job_id]
            
            # 机器约束：必须等待机器空闲
            # 检查该工序在机器上的正确位置
            machine_ops = machine_orders[machine]
            op_position_in_machine = -1
            for i, machine_op in enumerate(machine_ops):
                if machine_op.id == op.id:
                    op_position_in_machine = i
                    break
            
            # 计算机器的最早可用时间
            earliest_start_machine = machine_completion_times[machine]
            
            # 如果不是机器上的第一个工序，需要等待前面工序完成
            if op_position_in_machine > 0:
                # 找到在该机器上的直接前驱工序
                prev_machine_op = machine_ops[op_position_in_machine - 1]
                if prev_machine_op.id in start_times:
                    prev_finish_time = start_times[prev_machine_op.id] + prev_machine_op.get_processing_time(machine_assignment[prev_machine_op.id])
                    earliest_start_machine = max(earliest_start_machine, prev_finish_time)
            
            # 取两者的最大值作为最早开始时间
            earliest_start = max(earliest_start_job, earliest_start_machine)
            
            # 获取加工时间（不再有优化）
            processing_time = op.get_processing_time(machine_assignment[op.id])
            
            # 设置开始时间和完成时间
            start_times[op.id] = earliest_start
            completion_time = earliest_start + processing_time
            
            # 更新工件完成时间（该工件的最新工序完成时间）
            job_completion_times[job_id] = completion_time
            
            # 更新机器完成时间
            machine_completion_times[machine] = completion_time
        
        makespan = max(machine_completion_times.values()) if machine_completion_times else 0
        
        return makespan, start_times
    
    def _get_job_constraint_time(self, operation: FlexibleOperation, 
                                start_times: Dict[str, int], 
                                machine_assignment: Dict[str, str]) -> int:
        """
        获取工件内工序约束的最早开始时间
        """
        job_id = operation.job_id
        operation_id = operation.operation_id
        
        # 找到同一工件的所有工序
        job_operations = []
        for op in self.graph.operations:
            if op.job_id == job_id:
                job_operations.append(op)
        
        # 按工序ID排序
        job_operations.sort(key=lambda x: x.operation_id)
        
        # 找到当前工序的位置
        current_index = -1
        for i, op in enumerate(job_operations):
            if op.operation_id == operation_id:
                current_index = i
                break
        
        # 如果是第一个工序，无前序约束
        if current_index <= 0:
            return 0
        
        # 计算前序工序的完成时间
        max_predecessor_finish = 0
        for i in range(current_index):
            pred_op = job_operations[i]
            if pred_op.id in start_times:
                pred_start = start_times[pred_op.id]
                pred_duration = pred_op.get_processing_time(machine_assignment[pred_op.id])
                pred_finish = pred_start + pred_duration
                max_predecessor_finish = max(max_predecessor_finish, pred_finish)
        
        return max_predecessor_finish
    
    def _initialize_population(self, size: int) -> List[Tuple[Dict[str, str], Dict[str, List]]]:
        """初始化种群"""
        population = []
        
        for _ in range(size):
            # 随机机器分配
            machine_assignment = {}
            for op in self.schedule.operations:
                available_machines = op.get_available_machines()
                machine_assignment[op.id] = random.choice(available_machines)
            
            # 基于机器分配生成工序顺序
            machine_orders = self._generate_random_order(machine_assignment)
            
            population.append((machine_assignment, machine_orders))
        
        return population
    
    def _generate_random_order(self, machine_assignment: Dict[str, str]) -> Dict[str, List]:
        """基于机器分配生成随机工序顺序，严格保持同一工件内工序的顺序约束"""
        machine_operations = defaultdict(list)
        
        # 将工序分配到对应机器
        for op in self.schedule.operations:
            machine = machine_assignment[op.id]
            machine_operations[machine].append(op)
        
        # 对每台机器上的工序进行排序和优化
        for machine in machine_operations:
            operations = machine_operations[machine]
            
            # 按工件ID和工序ID严格排序，确保同一工件的工序按序号排列
            operations.sort(key=lambda x: (x.job_id, x.operation_id))
            
            # 创建工件工序组
            job_operation_groups = defaultdict(list)
            for op in operations:
                job_operation_groups[op.job_id].append(op)
            
            # 确保每个工件内的工序保持顺序
            for job_id in job_operation_groups:
                job_operation_groups[job_id].sort(key=lambda x: x.operation_id)
            
            # 重新调度：采用轮转方式处理不同工件，但保持工件内顺序
            result = []
            job_indices = {job_id: 0 for job_id in job_operation_groups.keys()}
            
            while any(job_indices[job_id] < len(job_operation_groups[job_id]) 
                     for job_id in job_operation_groups):
                
                # 获取所有还有未调度工序的工件
                available_jobs = [job_id for job_id in job_operation_groups.keys()
                                if job_indices[job_id] < len(job_operation_groups[job_id])]
                
                if not available_jobs:
                    break
                
                # 随机选择一个工件进行调度
                selected_job = random.choice(available_jobs)
                
                # 添加该工件的下一个工序
                next_op = job_operation_groups[selected_job][job_indices[selected_job]]
                result.append(next_op)
                job_indices[selected_job] += 1
            
            machine_operations[machine] = result
        
        return dict(machine_operations)
    
    def _roulette_selection(self, population: List, fitness_scores: List[float]):
        """轮盘赌选择"""
        total_fitness = sum(fitness_scores)
        if total_fitness == 0:
            return random.choice(population)
        
        pick = random.uniform(0, total_fitness)
        current = 0
        for i, fitness in enumerate(fitness_scores):
            current += fitness
            if current >= pick:
                return population[i]
        return population[-1]
    
    def _crossover(self, parent1: Tuple, parent2: Tuple) -> Tuple[Tuple, Tuple]:
        """交叉操作 - 保持工件内工序顺序"""
        assignment1, orders1 = parent1
        assignment2, orders2 = parent2
        
        # 机器分配交叉
        child_assignment1 = assignment1.copy()
        child_assignment2 = assignment2.copy()
        
        # 随机选择一部分工序进行交换
        operations = list(assignment1.keys())
        crossover_point = len(operations) // 2
        random.shuffle(operations)
        
        for op_id in operations[:crossover_point]:
            child_assignment1[op_id] = assignment2[op_id]
            child_assignment2[op_id] = assignment1[op_id]
        
        # 基于新的机器分配重新生成工序顺序（保持约束）
        child_orders1 = self._generate_random_order(child_assignment1)
        child_orders2 = self._generate_random_order(child_assignment2)
        
        return (child_assignment1, child_orders1), (child_assignment2, child_orders2)
    
    def _mutate(self, individual: Tuple) -> Tuple:
        """变异操作 - 保持工件内工序顺序"""
        assignment, orders = individual
        new_assignment = assignment.copy()
        
        # 随机改变几个工序的机器分配
        operations = list(assignment.keys())
        num_mutations = max(1, len(operations) // 20)  # 减少变异强度
        
        for _ in range(num_mutations):
            op_id = random.choice(operations)
            # 找到对应的工序对象
            operation = None
            for op in self.schedule.operations:
                if op.id == op_id:
                    operation = op
                    break
            
            if operation:
                available_machines = operation.get_available_machines()
                if len(available_machines) > 1:  # 只有多个选择时才变异
                    current_machine = new_assignment[op_id]
                    available_machines = [m for m in available_machines if m != current_machine]
                    if available_machines:
                        new_assignment[op_id] = random.choice(available_machines)
        
        # 重新生成工序顺序（保持约束）
        new_orders = self._generate_random_order(new_assignment)
        
        return (new_assignment, new_orders)

    def genetic_algorithm(self, population_size: int = 100, generations: int = 500,
                         mutation_rate: float = 0.1, crossover_rate: float = 0.8) -> Tuple[int, Dict, Dict]:
        """
        遗传算法求解FJSP - 增加负载平衡优化
        """
        print(f"开始遗传算法求解 (种群大小: {population_size}, 代数: {generations})")
        
        # 初始化种群
        population = self._initialize_population(population_size)
        
        best_makespan = float('inf')
        best_solution = None
        best_orders = None
        generation_bests = []
        
        for generation in range(generations):
            # 评估种群适应度
            fitness_scores = []
            for individual in population:
                machine_assignment, machine_orders = individual
                makespan, _ = self.calculate_makespan(machine_assignment, machine_orders)
                
                # 计算负载平衡因子
                machine_loads = defaultdict(int)
                for machine, operations in machine_orders.items():
                    for op in operations:
                        machine_loads[machine] += op.get_processing_time(machine_assignment[op.id])
                
                if machine_loads:
                    avg_load = sum(machine_loads.values()) / len(machine_loads)
                    load_variance = sum((load - avg_load)**2 for load in machine_loads.values()) / len(machine_loads)
                    balance_factor = 1 / (1 + load_variance / (avg_load**2 + 1))
                else:
                    balance_factor = 1
                
                # 综合适应度：考虑makespan和负载平衡
                fitness = (1.0 / (makespan + 1)) * (0.7 + 0.3 * balance_factor)
                fitness_scores.append(fitness)
                
                if makespan < best_makespan:
                    best_makespan = makespan
                    best_solution = machine_assignment.copy()
                    best_orders = {m: ops.copy() for m, ops in machine_orders.items()}
            
            generation_bests.append(best_makespan)
            
            # 选择、交叉、变异
            new_population = []
            
            # 保留最优个体（精英策略）
            best_idx = fitness_scores.index(max(fitness_scores))
            new_population.append(population[best_idx])
            
            # 生成新个体
            while len(new_population) < population_size:
                # 轮盘赌选择
                parent1 = self._roulette_selection(population, fitness_scores)
                parent2 = self._roulette_selection(population, fitness_scores)
                
                # 交叉
                if random.random() < crossover_rate:
                    child1, child2 = self._crossover(parent1, parent2)
                else:
                    child1, child2 = parent1, parent2
                
                # 变异
                if random.random() < mutation_rate:
                    child1 = self._mutate(child1)
                if random.random() < mutation_rate:
                    child2 = self._mutate(child2)
                
                new_population.extend([child1, child2])
            
            population = new_population[:population_size]
            
            print(f"第 {generation} 代，最佳makespan: {best_makespan}")
        
        print(f"遗传算法完成，最佳makespan: {best_makespan}")
        
        # 绘制收敛曲线
        self._plot_convergence(generation_bests)
        
        return best_makespan, best_solution, best_orders
    
    def _plot_convergence(self, generation_bests: List[int]):
        """绘制算法收敛曲线"""
        plt.figure(figsize=(10, 6))
        plt.plot(generation_bests, 'b-', linewidth=2, alpha=0.8)
        plt.xlabel('Generation', fontsize=12)
        plt.ylabel('Best Makespan', fontsize=12)
        plt.title('遗传算法收敛曲线', fontsize=14)
        plt.grid(True, alpha=0.3)
        plt.tight_layout()
        plt.show()
    


def validate_job_data(job_data: List[Tuple[List[int], List[int]]], num_jobs: int, num_machines: int) -> List[Job]:
    """
    验证工件数据并创建工件对象
    
    Args:
        job_data: 工件数据列表
        num_jobs: 工件数量
        num_machines: 机器数量
        
    Returns:
        List[Job]: 验证后的工件对象列表
    """
    print("正在验证数据完整性...")
    
    # 检查数据合理性
    for i, (machines, times) in enumerate(job_data):
        if len(machines) != len(times):
            raise ValueError(f"工件 {i+1} 的机器和时间数据长度不匹配")
        
        if any(t <= 0 for t in times):
            print(f"警告: 工件 {i+1} 包含非正数加工时间")
        
        if any(m < 1 for m in machines):
            print(f"警告: 工件 {i+1} 包含无效机器编号")
    
    # 创建工件对象
    jobs = []
    for i, (machines, times) in enumerate(job_data):
        try:
            job = Job(i + 1, machines, times)
            jobs.append(job)
            if i < 3:
                print(f"工件 {i+1}: 机器序列长度={len(machines)}, 加工时间范围=[{min(times)}, {max(times)}]")
        except Exception as e:
            raise ValueError(f"创建工件 {i+1} 时发生错误: {e}")
    
    return jobs

def build_disjunctive_graph_and_solver(jobs: List[Job]) -> Tuple[DisjunctiveGraph, ScheduleSolver]:
    """
    构建析取图和求解器
    
    Args:
        jobs: 工件列表
        
    Returns:
        Tuple[DisjunctiveGraph, ScheduleSolver]: 析取图和求解器
    """
    try:
        graph = DisjunctiveGraph(jobs)
        solver = ScheduleSolver(graph)
        
        print(f"\n析取图构建完成:")
        print(f"- 工序总数: {len(graph.operations)}")
        print(f"- 连接弧数量: {len(graph.conjunctive_arcs)}")
        print(f"- 析取弧数量: {len(graph.disjunctive_arcs)}")
        
        # 验证机器操作分布
        print(f"- 各机器上的工序数量:")
        for machine in sorted(graph.machine_operations.keys()):
            ops_count = len(graph.machine_operations[machine])
            print(f"  机器 M{machine}: {ops_count} 个工序")
            
        return graph, solver
        
    except Exception as e:
        raise ValueError(f"构建析取图时发生错误: {e}")

def solve_and_visualize(solver: ScheduleSolver, dataset_name: str, num_jobs: int, max_iterations: int, max_attempts: int) -> Dict[str, Tuple]:
    """
    求解调度问题并可视化结果
    
    Args:
        solver: 调度求解器
        dataset_name: 数据集名称
        num_jobs: 工件数量
        max_iterations: 最大迭代次数
        max_attempts: 最大尝试次数
        
    Returns:
        Dict[str, Tuple]: 求解结果
    """
    print(f"\n运行随机搜索算法 (最大迭代次数: {max_iterations})...")
    
    try:
        start_time = time.time()
        best_makespan_rs, best_orders_rs = solver.random_search(max_iterations=max_iterations, max_attempts=max_attempts)
        rs_time = time.time() - start_time
        
        if best_makespan_rs != float('inf'):
            print(f"随机搜索完成: Makespan = {best_makespan_rs}, 时间 = {rs_time:.3f}s")
            
            # 生成甘特图
            print(f"\n正在为 {dataset_name} 生成甘特图...")
            try:
                _, start_times = solver.calculate_makespan(best_orders_rs)
                visualize_schedule(best_orders_rs, start_times, 
                                    f"{dataset_name} - 随机搜索 (Makespan = {best_makespan_rs})")
                
                # 显示调度详情
                if num_jobs <= 10:  # 只对小规模问题显示详细信息
                    print_schedule_details(best_orders_rs, start_times, best_makespan_rs)
                    
            except Exception as e:
                print(f"生成甘特图时出错: {e}")
                
            return {'随机搜索': (best_makespan_rs, rs_time)}
        else:
            print(f"随机搜索未找到有效解")
            return {'随机搜索': (None, None)}
            
    except Exception as e:
        print(f"随机搜索执行出错: {e}")
        return {'随机搜索': (None, None)}

def read_real_case_data(filename: str) -> Tuple[List[FlexibleJob], Dict[str, str]]:
    """
    读取实际案例Excel数据
    
    Args:
        filename: Excel文件路径
        
    Returns:
        (jobs, machine_mapping): 柔性工件列表和机器映射表
    """

    # 使用openpyxl读取Excel文件
    workbook = openpyxl.load_workbook(filename)
    worksheet = workbook.active
    
    print(f"读取Excel文件: {filename}")
    print(f"工作表名称: {worksheet.title}")
    print(f"数据范围: {worksheet.max_row} 行 x {worksheet.max_column} 列")
    
    # 解析数据结构
    jobs = []
    machine_mapping = {}
    
    # 从第5行开始读取数据
    current_job_id = 1
    current_job_operations = []
    current_quantity = 1
    
    # 机器列名映射 (F列开始对应所有机器，E列是工序序号)
    machine_columns = MACHINE_LIST
    part_name = ""
    for row in range(6, worksheet.max_row + 1):  # 从第6行开始

        # 读取各列数据
        cur_part_name = worksheet.cell(row=row, column=2).value  # B列：零件名称
        
        cur_quantity = worksheet.cell(row=row, column=3).value   # C列：数量
        feature = worksheet.cell(row=row, column=4).value    # D列：特征
        operation_seq = worksheet.cell(row=row, column=5).value  # E列：工序序号
        feature = str(feature).strip() if feature else ""
        
        # 处理数量
        quantity = int(float(str(cur_quantity).strip())) if cur_quantity else quantity
        
        # 处理工序序号
        operation_seq = int(float(str(operation_seq).strip())) if operation_seq else 0

        # 检查是否开始新工件（B列出现新值）
        if cur_part_name is not None and part_name != cur_part_name:
            
            # 创建前一个工件（根据数量创建多个相同工件）
            if current_job_operations:
                for i in range(current_quantity):
                    # 为每个工件创建独立的工序副本，使用不同的job_id
                    job_operations = []
                    for op in current_job_operations:
                        new_op = FlexibleOperation(
                            current_job_id, 
                            op.operation_id, 
                            op.machine_options.copy(), 
                            part_name
                        )
                        job_operations.append(new_op)
                    
                    job = FlexibleJob(current_job_id, job_operations, part_name)
                    jobs.append(job)
                    current_job_id += 1
            part_name = str(cur_part_name).strip() if cur_part_name else part_name
            current_job_operations = []

        current_quantity = quantity
        
        # 解析机器选项和时间（从F列开始）
        machine_options = []
        
        # 读取F列到后续列的机器时间数据
        for i, machine_name in enumerate(machine_columns):
            col_idx = 6 + i  # F列开始（索引6）
            cell_value = worksheet.cell(row=row, column=col_idx).value
            
            if cell_value is not None and str(cell_value).strip() != '':
                try:
                    time_val = int(float(str(cell_value).strip()))
                    if time_val > 0:  # 只添加正数时间
                        machine_options.append((machine_name, time_val))
                except (ValueError, TypeError):
                    continue
        print(f"工序 {operation_seq} - 零件 {part_name}, 数量 {current_quantity}, 机器选项: {machine_options}")
        # 如果有有效的机器选项，创建工序
        if machine_options and operation_seq > 0:
            # 使用临时的job_id创建模板工序
            operation = FlexibleOperation(0, operation_seq, machine_options, part_name)
            current_job_operations.append(operation)
    
    # 添加最后一个工件（根据数量创建多个相同工件）
    if current_job_operations:
        for i in range(current_quantity):
            # 为每个工件创建独立的工序副本，使用不同的job_id
            job_operations = []
            for op in current_job_operations:
                new_op = FlexibleOperation(
                    current_job_id, 
                    op.operation_id, 
                    op.machine_options.copy(), 
                    part_name
                )
                job_operations.append(new_op)
            
            job = FlexibleJob(current_job_id, job_operations, part_name)
            jobs.append(job)
            current_job_id += 1

    # 创建机器映射
    all_machines = set()
    for job in jobs:
        for op in job.operations:
            for machine, _ in op.machine_options:
                all_machines.add(machine)
    
    machine_mapping = {machine: machine for machine in sorted(all_machines)}
    
    print(f"\n解析完成:")
    print(f"工件数量: {len(jobs)}")
    print(f"机器数量: {len(machine_mapping)}")
    print(f"机器列表: {sorted(machine_mapping.keys())}")
    
    workbook.close()
    return jobs, machine_mapping


def visualize_flexible_schedule(machine_assignment: Dict[str, str],
                               machine_orders: Dict[str, List[FlexibleOperation]], 
                               start_times: Dict[str, int], 
                               title: str = "柔性作业车间调度甘特图"):
    """
    可视化柔性调度方案的甘特图 - 修复显示问题
    """
    if not machine_orders or not start_times:
        print("无效的调度数据，无法生成甘特图")
        return
        
    fig, ax = plt.subplots(figsize=(20, 12))
    
    # 扩展颜色映射
    colors = [
        '#FF6B6B', '#4ECDC4', '#45B7D1', '#96CEB4', '#FFEAA7', 
        '#DDA0DD', '#98D8C8', '#F7DC6F', '#BB8FCE', '#85C1E9',
        '#F8C471', '#82E0AA', '#F1948A', '#85C1E9', '#D2B4DE',
        '#AED6F1', '#A3E4D7', '#D5DBDB', '#FADBD8', '#D1F2EB',
        '#FCF3CF', '#EBDEF0', '#EAF2F8', '#E8F8F5', '#FDF2E9'
    ]
    
    # 为不同产品分配颜色
    product_colors = {}
    job_colors = {}
    
    machines = sorted(machine_orders.keys())
    y_positions = {machine: i for i, machine in enumerate(machines)}
    
    # 计算最大时间
    max_time = 0
    for ops in machine_orders.values():
        for op in ops:
            if op.id in start_times:
                finish_time = start_times[op.id] + op.get_processing_time(machine_assignment[op.id])
                max_time = max(max_time, finish_time)
    
    print(f"甘特图信息:")
    print(f"  机器数量: {len(machines)}")
    print(f"  总工序数: {sum(len(ops) for ops in machine_orders.values())}")
    print(f"  最大时间: {max_time}")
    
    # 绘制甘特图
    for machine, operations in machine_orders.items():
        y_pos = y_positions[machine]
        print(f"  机器 {machine}: {len(operations)} 个工序")
        
        for i, op in enumerate(operations):
            # 为产品类型分配颜色
            if op.product_type not in product_colors:
                product_colors[op.product_type] = colors[len(product_colors) % len(colors)]
            
            # 为工件分配相同产品的颜色
            if op.job_id not in job_colors:
                base_color = product_colors[op.product_type]
                job_colors[op.job_id] = base_color
            
            if op.id in start_times:
                start_time = start_times[op.id]
                duration = op.get_processing_time(machine_assignment[op.id])
                
                print(f"    工序 {op.id}: 开始={start_time}, 持续={duration}, 结束={start_time + duration}")
                
                # 绘制工序矩形
                rect = patches.Rectangle((start_time, y_pos - 0.35), duration, 0.7,
                                       linewidth=1.5, edgecolor='white', 
                                       facecolor=job_colors[op.job_id], alpha=0.8)
                ax.add_patch(rect)
                
                # 添加工序标签
                if duration > max_time * 0.01:  # 降低显示文字的阈值
                    ax.text(start_time + duration/2, y_pos, f'J{op.job_id}O{op.operation_id}',
                           ha='center', va='center', fontsize=9, fontweight='bold', color='black')
    
    # 设置图表属性
    ax.set_xlabel('Time (时间单位)', fontsize=16, fontweight='bold')
    ax.set_ylabel('Machine (机器)', fontsize=16, fontweight='bold')
    ax.set_title(title, fontsize=18, fontweight='bold', pad=20)
    
    # 设置y轴
    ax.set_yticks(range(len(machines)))
    ax.set_yticklabels([f'{m}' for m in machines], fontsize=12)
    ax.set_ylim(-0.5, len(machines) - 0.5)
    
    # 设置x轴
    if max_time > 0:
        ax.set_xlim(0, max_time + max_time * 0.05)
        # 添加时间刻度
        time_ticks = range(0, int(max_time) + 1, max(1, int(max_time) // 20))
        ax.set_xticks(time_ticks)
    ax.tick_params(axis='x', labelsize=12)
    ax.tick_params(axis='y', labelsize=12)
    
    # 添加网格
    ax.grid(True, alpha=0.3, linestyle='-', linewidth=0.5)
    ax.set_axisbelow(True)
    
    # 添加产品图例
    legend_elements = []
    for product_type in sorted(product_colors.keys()):
        legend_elements.append(
            patches.Patch(facecolor=product_colors[product_type], 
                         label=f'{product_type}', alpha=0.8,
                         edgecolor='white', linewidth=1)
        )
    
    if legend_elements:
        ax.legend(handles=legend_elements, loc='center left', 
                 bbox_to_anchor=(1.02, 0.5), fontsize=12,
                 title='产品类型', title_fontsize=14)
    
    # 添加makespan标记线
    if max_time > 0:
        ax.axvline(x=max_time, color='red', linestyle='--', linewidth=3, alpha=0.8)
        ax.text(max_time + max_time * 0.01, len(machines) * 0.95, 
               f'Makespan = {max_time}', 
               fontsize=14, fontweight='bold', color='red',
               bbox=dict(boxstyle="round,pad=0.3", facecolor="white", alpha=0.8))
    
    # 添加机器利用率信息
    machine_utilization = {}
    for machine in machines:
        total_work_time = 0
        if machine in machine_orders:
            for op in machine_orders[machine]:
                if op.id in start_times:
                    total_work_time += op.get_processing_time(machine_assignment[op.id])
        utilization = (total_work_time / max_time * 100) if max_time > 0 else 0
        machine_utilization[machine] = utilization
    
    # 在图表上显示利用率
    for i, machine in enumerate(machines):
        util = machine_utilization[machine]
        ax.text(-max_time * 0.02, i, f'{util:.1f}%', 
               ha='right', va='center', fontsize=10, 
               bbox=dict(boxstyle="round,pad=0.2", facecolor="lightblue", alpha=0.7))
    
    ax.set_facecolor('#FAFAFA')
    plt.tight_layout()
    plt.show()
    
    # 打印详细的调度统计
    print(f"\n机器利用率统计:")
    for machine in sorted(machines):
        util = machine_utilization[machine]
        work_time = sum(op.get_processing_time(machine_assignment[op.id]) 
                       for op in machine_orders.get(machine, []) 
                       if op.id in start_times)
        print(f"  {machine}: 工作时间={work_time}, 利用率={util:.1f}%")

def save_solution_to_file(machine_assignment: Dict[str, str],
                         machine_orders: Dict[str, List[FlexibleOperation]], 
                         start_times: Dict[str, int],
                         flexible_jobs: List[FlexibleJob],
                         filename: str = "realcase_sol.txt"):
    """
    将调度方案保存到文件中，按指定格式输出
    
    Args:
        machine_assignment: 机器分配方案
        machine_orders: 机器工序顺序
        start_times: 工序开始时间
        flexible_jobs: 柔性工件列表
        filename: 输出文件名
    """
    print(f"\n正在保存调度方案到文件: {filename}")
    
    # 创建解决方案记录列表
    solution_records = []
    
    # 创建产品类型到零件编号的映射
    product_to_number = {}
    unique_products = set()
    for job in flexible_jobs:
        unique_products.add(job.product_type)
    
    # 按产品类型排序并分配编号
    for i, product_type in enumerate(sorted(unique_products), 1):
        product_to_number[product_type] = i
    
    # 计算工件数量和机器数量
    num_jobs = len(flexible_jobs)
    num_machines = len(set(machine_assignment.values()))
    
    # 遍历每台机器的工序安排
    for machine_name, operations in machine_orders.items():
        for op in operations:
            if op.id in start_times:
                start_time = start_times[op.id]
                processing_time = op.get_processing_time(machine_assignment[op.id])
                end_time = start_time + processing_time
                
                # 获取零件编号（基于产品类型）
                part_number = op.job_id
                
                solution_records.append({
                    'equipment': machine_name,
                    'part_number': part_number,
                    'operation_number': op.operation_id,
                    'start_time': start_time,
                    'end_time': end_time,
                    'product_type': op.product_type,
                    'job_id': op.job_id
                })
    
    # 按设备编号和开始时间排序
    solution_records.sort(key=lambda x: (x['equipment'], x['start_time']))
    
    # 写入文件
    with open(filename, 'w', encoding='utf-8') as f:
        # 写入工件数量和机器数量信息
        f.write(f"Nb of jobs {num_jobs} Nb of Machines {num_machines}\n")
        
        # 写入表头
        f.write("Solution\n")
        f.write(f"{'设备编号':<12} {'零件编号':<12} {'工序标号':<12} {'开始时间':<12} {'结束时间':<12}\n")
        
        # 写入调度记录
        current_equipment = None
        for record in solution_records:
            equipment = record['equipment']
            part_num = record['part_number']
            operation_num = record['operation_number']
            start_time = record['start_time']
            end_time = record['end_time']

            f.write(f"{equipment:<12} {part_num:<12} {operation_num:<12} {start_time:<12} {end_time:<12}\n")
        


def solve_real_case(max_iterations: int = 5000):
    """求解实际案例"""
    print("开始处理实际案例数据...")
    
    # 读取Excel数据
    filename = "data/real_case.xlsx"
    flexible_jobs, machine_mapping = read_real_case_data(filename)
    
    if not flexible_jobs:
        print("未能成功读取数据，请检查Excel文件格式")
        return
    
    graph = FlexibleSchedule(flexible_jobs)
    solver = FlexibleScheduleSolver(graph)
    
    print(f"工序总数: {len(graph.operations)}")
    print(f"机器总数: {len(graph.all_machines)}")
    print(f"机器列表: {sorted(graph.all_machines)}")

    
    # 显示产品类型统计
    product_stats = defaultdict(int)
    for job in flexible_jobs:
        product_stats[job.product_type] += 1
    
    print(f"\n产品类型统计:")
    for product, count in sorted(product_stats.items()):
        print(f"  {product}: {count} 个工件")
    
    # 运行遗传算法
    print(f"\n开始求解柔性作业车间调度问题...")
    start_time = time.time()
    
    best_makespan, best_assignment, best_orders = solver.genetic_algorithm(
        population_size=100, 
        generations=max_iterations,
        mutation_rate=0.12,
        crossover_rate=0.88
    )
    
    solve_time = time.time() - start_time
    
    if best_makespan != float('inf'):
        print(f"\n求解完成!")
        print(f"最优makespan: {best_makespan}")
        print(f"求解时间: {solve_time:.3f}秒")
        
        # 计算详细调度信息
        _, start_times = solver.calculate_makespan(best_assignment, best_orders)
        
        # 保存调度方案到文件
        save_solution_to_file(best_assignment, best_orders, start_times, flexible_jobs, "realcase_sol.txt")
        
        # 生成甘特图
        visualize_flexible_schedule(
            best_assignment, best_orders, start_times,
            f"柔性作业车间调度结果 (Makespan = {best_makespan})"
        )

def visualize_schedule(machine_orders: Dict[int, List[Operation]], 
                      start_times: Dict[str, int], 
                      title: str = "作业车间调度甘特图"):
    """
    可视化调度方案的甘特图
    
    Args:
        machine_orders: 每台机器的工序顺序
        start_times: 各工序的开始时间
        title: 图表标题
    """
    # 检查输入数据
    if not machine_orders or not start_times:
        print("无效的调度数据，无法生成甘特图")
        return
        
    fig, ax = plt.subplots(figsize=(14, 8))
    
    # 扩展颜色映射（为不同工件分配不同颜色）
    colors = [
        '#FF6B6B', '#4ECDC4', '#45B7D1', '#96CEB4', '#FFEAA7', 
        '#DDA0DD', '#98D8C8', '#F7DC6F', '#BB8FCE', '#85C1E9',
        '#F8C471', '#82E0AA', '#F1948A', '#85C1E9', '#D2B4DE',
        '#AED6F1', '#A3E4D7', '#D5DBDB', '#FADBD8', '#D1F2EB',
        '#FCF3CF', '#EBDEF0', '#EAF2F8', '#E8F8F5', '#FDF2E9'
    ]
    job_colors = {}
    
    machines = sorted(machine_orders.keys())
    y_positions = {machine: i for i, machine in enumerate(machines)}
    
    # 计算最大时间用于设置x轴
    max_time = max(start_times[op.id] + op.processing_time 
                  for ops in machine_orders.values() for op in ops)
    
    # 绘制甘特图
    for machine, operations in machine_orders.items():
        y_pos = y_positions[machine]
        
        for op in operations:
            job_id = op.job_id
            if job_id not in job_colors:
                job_colors[job_id] = colors[len(job_colors) % len(colors)]
            
            start_time = start_times[op.id]
            duration = op.processing_time
            
            # 绘制工序矩形 - 去掉文字标签
            rect = patches.Rectangle((start_time, y_pos - 0.35), duration, 0.7,
                                   linewidth=1.5, edgecolor='white', 
                                   facecolor=job_colors[job_id], alpha=0.8)
            ax.add_patch(rect)
    
    # 设置图表属性
    ax.set_xlabel('Time', fontsize=16, fontweight='bold')
    ax.set_ylabel('Machine', fontsize=16, fontweight='bold')
    ax.set_title(title, fontsize=18, fontweight='bold', pad=20)
    
    # 设置y轴
    ax.set_yticks(range(len(machines)))
    ax.set_yticklabels([f'M{m}' for m in machines], fontsize=14)
    ax.set_ylim(-0.5, len(machines) - 0.5)
    
    # 设置x轴
    ax.set_xlim(0, max_time + max_time * 0.05)  # 添加5%的边距
    ax.tick_params(axis='x', labelsize=12)
    ax.tick_params(axis='y', labelsize=12)
    
    # 添加网格
    ax.grid(True, alpha=0.3, linestyle='-', linewidth=0.5)
    ax.set_axisbelow(True)
    
    # 添加图例 - 显示工件信息
    legend_elements = []
    for job_id in sorted(job_colors.keys()):
        legend_elements.append(
            patches.Patch(facecolor=job_colors[job_id], 
                         label=f'Job {job_id}', alpha=0.8,
                         edgecolor='white', linewidth=1)
        )
    
    # 将图例放在图表右侧
    ax.legend(handles=legend_elements, loc='center left', 
             bbox_to_anchor=(1.02, 0.5), fontsize=12,
             title='Jobs', title_fontsize=14)
    
    # 添加makespan标记线
    ax.axvline(x=max_time, color='red', linestyle='--', linewidth=3, alpha=0.8)
    ax.text(max_time + max_time * 0.01, len(machines) * 0.95, 
           f'Makespan = {max_time}', 
           fontsize=14, fontweight='bold', color='red',
           bbox=dict(boxstyle="round,pad=0.3", facecolor="white", alpha=0.8))
    
    # 设置背景色
    ax.set_facecolor('#FAFAFA')
    
    plt.tight_layout()
    plt.show()

def print_schedule_details(machine_orders: Dict[int, List[Operation]], 
                          start_times: Dict[str, int], 
                          makespan: int):
    """
    打印调度方案的详细信息
    
    Args:
        machine_orders: 每台机器的工序顺序
        start_times: 各工序的开始时间
        makespan: 最大完工时间
    """
    print("\n" + "="*60)
    print("调度方案详细信息")
    print("="*60)
    
    for machine in sorted(machine_orders.keys()):
        print(f"\n机器 M{machine} 的加工顺序:")
        operations = machine_orders[machine]
        
        for i, op in enumerate(operations):
            start_time = start_times[op.id]
            finish_time = start_time + op.processing_time
            print(f"  {i+1}. {op.id} (工件J{op.job_id}) | "
                  f"开始时间: {start_time}, 结束时间: {finish_time}, "
                  f"加工时间: {op.processing_time}")
    
    print(f"\n总完工时间 (Makespan): {makespan}")
    print("="*60)


if __name__ == "__main__":
    # 设置中文字体支持
    plt.rcParams['font.sans-serif'] = ['SimHei', 'DejaVu Sans']
    plt.rcParams['axes.unicode_minus'] = False

    # 运行柔性作业车间调度实际案例
    solve_real_case(max_iterations=1)
